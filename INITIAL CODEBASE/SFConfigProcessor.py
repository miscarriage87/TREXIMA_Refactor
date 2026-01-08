from bs4 import BeautifulSoup, NavigableString
import os
import babel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Protection, colors, GradientFill, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, Text, ttk, PhotoImage, messagebox
from ttkthemes import ThemedTk
import requests
import pyodata
import easygui as gui
import sys
import time
import csv

# pylint: disable=locally-disabled, multiple-statements, fixme, line-too-long, no-member


appName = "TREXIMA"
versionNum = 3.0
appTitle = "{} - Translation Export & Import Accelerator".format(appName)
appUI = ThemedTk(theme='elegance')
appUI.title(appTitle)

appIconFileName = 'appicon.png'
appDir = os.path.dirname(__file__)
pathToThisApp = os.path.normpath(appDir)
stdSDMFileName = "EC-data-model.xml"
stdCSF_SDMFileName = "EC-CSF-for-succession-DM.xml"
stdCDMFileName = "EC-corporate-datamodel.xml"
stdCSF_CDMFileName = "EC-CSF-for-corporate-DM.xml"
KEYWORD_SAP_STANDARD = "Standard SAP"
sysDefaultLang = "en_US"

fullPathToAppIcon = os.path.join(pathToThisApp, appIconFileName)
photo = PhotoImage(file = fullPathToAppIcon)
appUI.iconphoto(False, photo)
#windowApTrans.wm_iconbitmap('ApTrans.ico')
appUI.geometry("1200x700")
appUI.state('zoomed') 
bgcolor="#066b8d"
appUI.config(background=bgcolor)
sfOdataService = None
soups = []
soupsDict = {}
fileSaveDir = "C:\\"
askToChooseDMLangs = True
excludeCustomLabelStrdFields = True 
excelFileName = None
importFilePath = None
translationsWb = None
xmlFiles = []
translatableTags = []
isPMGMIncluded = False
isSDMIncluded = False
labelKeysFile = None
labelKeysDict = {} 
labelKeysFileHeaders = []
activeCountries = []
childChar = " ↪ "
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
thick = Side(border_style="thick", color="000000")
headerCellBorder = Border(top=thick, left=thick, right=thick, bottom=thick)
boldFont = Font(color='00000000', bold=True, size=12)

sheetNameForPM = "Performance_Review_Templates"
sheetNameForGM = "Goal&Development_Plan_Templates"
sheetNameForPL = "Picklists"

employeeProfileTags = ["standard-element", "background-element", "userinfo-element", 
                    "data-field", "rating-field", "tab-element", "view-template", "edit-template"]
tagsToBeIgnored = ["tab-element", "view-template", "edit-template","fm-competency","permission"]

fullPathToBgImg = os.path.join(pathToThisApp, "bg.png")
background_image = PhotoImage(file = fullPathToBgImg)
background_label = tk.Label(appUI, image=background_image)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

pathToDoneIcon = os.path.join(pathToThisApp, 'done.png')
doneIcon = PhotoImage(file = pathToDoneIcon)

isExportAction = False
fileOpenBtn = None
connectApiBtn = None
exportExecuteBtn = None
importExecuteBtn = None

picklistIds = []
picklistReferences = []

'''
canvas = tk.Canvas(windowApTrans, height=700, width=1200, bg=bgcolor)
canvas.place(relx=0, rely=0, relheight=1, relwidth=1)
'''
# Progress bar widget 
s = ttk.Style(appUI)
# add the label to the progressbar style
s.layout("LabeledProgressbar",
         [('LabeledProgressbar.trough',
           {'children': [('LabeledProgressbar.pbar',
                          {'side': 'left', 'sticky': 'ns'}),
                         ("LabeledProgressbar.label",
                          {"sticky": ""})],
           'sticky': 'nswe'})])
TROUGH_COLOR = 'white'
BAR_COLOR = '#90ee90'

progress = ttk.Progressbar(appUI, orient = tk.HORIZONTAL, 
              length = 800, mode = 'determinate', style="LabeledProgressbar")
progressLog = ""


def logTheProgress(progressPercent, progressMsg):
    global progressLog

    timestamp = time.strftime("%a_%d%b_%Y_%Hh%Mm%Ss", time.localtime())
    progressLog += "{}: {}\n\n".format(timestamp, progressMsg)

    progress['value'] = progressPercent
    s.configure("LabeledProgressbar", text=progressMsg,
                troughcolor=TROUGH_COLOR, bordercolor=TROUGH_COLOR, background=BAR_COLOR, lightcolor=BAR_COLOR, darkcolor=BAR_COLOR)
    appUI.update()


def resetExecutionProgress():
    global fileOpenBtn
    global connectApiBtn
    global exportExecuteBtn
    global importExecuteBtn
    global importFilePath
    global labelKeysFile

    progressMsg = ""
    if fileOpenBtn != None:
        #fileOpenBtn["state"] = "disabled"
        #fileOpenBtn["image"] = ""
        fileOpenBtn.place_forget()
    if connectApiBtn != None:
        #connectApiBtn["state"] = "disabled"
        #connectApiBtn["image"] = ""
        connectApiBtn.place_forget()
    
    if isExportAction:
        if exportExecuteBtn != None:
            #exportExecuteBtn["state"] = "disabled"
            #exportExecuteBtn["image"] = ""
            exportExecuteBtn.place_forget()
        progressMsg = "The 'in-progress' Translations Export has been reset!"
    else:
        if importExecuteBtn != None:
            #importExecuteBtn["state"] = "disabled"
            #importExecuteBtn["image"] = ""
            importExecuteBtn.place_forget()
        importFilePath = None
        progressMsg = "The 'in-progress' Translations Import has been reset!"
    
    labelKeysFile = None
    
    logTheProgress(0, progressMsg)

def showMessage(title, text, option=None):
    root = tk.Tk()
    root.withdraw()
    retVal = None
    if title == None:
        title = "USEFUL INFORMATION"
    mb = messagebox
    if option == mb.YESNO:
        retVal = mb.askyesno(title, text)
    elif option == mb.RETRYCANCEL:
        retVal = mb.askretrycancel(title, text)
    elif option == mb.OKCANCEL:
        retVal = mb.askokcancel(title, text)
    elif option == mb.INFO or option == None:
        mb.showinfo(title, text)
    elif option == mb.WARNING:
        mb.showwarning(title, text)
    elif option == mb.ERROR:
        mb.showerror(title, text)
    root.destroy()
    return retVal

def getValuesAsMatrixFromCSV(pathToCSVFile):
    valuesMatrix = []

    with open(pathToCSVFile, newline='', encoding='utf-8') as csvfile:
        valuesListReader = csv.reader(csvfile)
        for row in valuesListReader:
            valuesMatrix.append(row)

    return valuesMatrix

def modifyCSVWith(pathToCSVFile, valuesMatrix, saveAsNewFile =True):
    updatedSuccessfully = False
    pathToModifiedCSV = pathToCSVFile;
    if saveAsNewFile:
        fileName = pathToCSVFile[pathToCSVFile.rfind("\\")+1:]
        fileSaveDir = pathToCSVFile[:pathToCSVFile.rfind("\\")]
        timestampSuffix = time.strftime("%a_%d%b_%Y_%Hh%Mm%Ss", time.localtime())
        pathToModifiedCSV = "{}{}Modified_{}_{}".format(fileSaveDir,"\\",timestampSuffix, fileName)

    with open(pathToModifiedCSV, 'w', newline='', encoding="utf-8") as csvFileToUpdate:
        writer = csv.DictWriter(csvFileToUpdate, valuesMatrix[0])
        writer.writeheader()
        writer.writerows(valuesMatrix[1:])

    return updatedSuccessfully


def getSFSpecificName(soup, isStandardDM = False):
    global isPMGMIncluded
    global labelKeysFile
    global labelKeysDict
    global labelKeysFileHeaders
    global activeCountries
    global isSDMIncluded
    specialName = None
    if soup.find("succession-data-model") != None and soup.find("hris-element") != None:
        
        if not isSDMIncluded:
            specialName = "SFEC Succession Data Model"
            isSDMIncluded = True
        else:
            specialName = "SFEC Succession Data Model"
    elif soup.find("succession-data-model") != None: 
        specialName = "SF Succession Data Model"
        isSDMIncluded = True
    elif soup.find("country-specific-fields") != None and soup.find("format-group") != None:
        specialName = "SFEC CSF Succession Data Model" 
        isSDMIncluded = True
        if len(activeCountries) == 0:
            continueToFile = showMessage("INFO about choice on the next prompted screen",
                            ("On the next prompted screen, browse to a CSV file that has code of " 
                            "countries which are enabled in the SF instance. Make sure that the " 
                            "1st column in the file contains 3 character country code in upper case. " 
                            "If you don't specify it, the export would include all countries from "
                            "the CSF data models. Click 'OK' to proceed once ready!"),
                            messagebox.OKCANCEL)
            if continueToFile:
                countryListFile = gui.fileopenbox('Browse to the Country List (csv) file.',
                            "Open Country List CSV File",fileSaveDir+"\\*.csv",["*.csv"])
                with open(countryListFile, newline='', encoding='utf-8') as csvfile:
                    countryListReader = csv.reader(csvfile)
                    for row in countryListReader:
                        firstCellValue = row[0]
                        if firstCellValue != None and len(firstCellValue)==3:
                            activeCountries.append(firstCellValue)
    elif soup.find("country-specific-fields") != None:
        specialName = "SFEC CSF Corporate Data Model"
    elif soup.find("corporate-data-model") != None:
        specialName = "SFEC Corporate Data Model"
    # Add config for PMGM Templates
    elif soup.find("sf-form") != None and soup.find("sf-pmreview") != None:
        specialName = "PM Form Template"
        isPMGMIncluded = True
        
        browseAgain = False
        if labelKeysFile != None:
            browseAgain = showMessage("Browse to FormLabelKeys file", 
            ("File '{}' has already been uploaded to get labels for PM Form " 
            "Template. Do you want to upload new 'FormLabelKeys' (csv) file? Click 'Conitinue' " 
            "to browse to a different 'FormLabelKeys' file. 'Cancel' to continue with already uploaded file.")
            .format(labelKeysFile),
            messagebox.OKCANCEL
            )
        else:
            browseAgain = True
        
        if browseAgain:
            showMessage('Useful Info for the Next Step',
                        ("On the next prompted screen, browse to the 'FormLabelKeys' (csv) file. " 
                        "This is an important and required file for extracting translations " 
                        "from PM Form Template for tags with attribute 'msgKey'. " 
                        "If you don't have it, please export it from the instance using the "
                        "tool 'Manage Form Label Translations' and then click " 
                        "'Export Form Label Translations'.\n\nClick 'OK' to proceed once ready!"),
                        messagebox.OKCANCEL)
            labelKeysFile = gui.fileopenbox('Browse to the "FormLabelKeys" (csv) file.',
                        "Open FormLabelKeys File",fileSaveDir+"\\*.csv",["*.csv"])
            with open(labelKeysFile, newline='', encoding='utf-8') as csvfile:
                labelKeysDictReader = csv.DictReader(csvfile)
                labelKeysFileHeaders = labelKeysDictReader.fieldnames
                row = None
                for row in labelKeysDictReader:
                    labelKeysDict[row['label_key']] = row
    elif soup.find("obj-plan-template") != None:
        specialName = "{} ({})".format(getDefaultTitle(soup.find("obj-plan-name")),
                                soup.find("obj-plan-id").string)
        isPMGMIncluded = True
    
    if isStandardDM:
        specialName = "{} {}".format(KEYWORD_SAP_STANDARD, specialName)

    return specialName

def getModuleSpecificName(parentTag):
    parentTagName = parentTag.name
    sectionName = parentTagName
    
    if parentTagName == "obj-plan-template":
        sectionName = "General Settings"
    elif (parentTagName == "text-replacement"):
        sectionName = "{} (for={})".format(getReadableName(parentTagName), parentTag.get("for"))
    elif parentTagName == "permission":
        sectionName = "Permission (for={})".format(parentTag.get("for"))
    elif parentTagName == "field-permission":
        sectionName = "Field Permission (type={})".format(parentTag.get("type"))
    elif parentTagName == "field-definition":
        sectionName = "Field Definition (id={})".format(parentTag.get("id"))
    elif parentTagName == "table-column":
        sectionName = childChar + "Table Column (id={})".format(parentTag.get("id"))
    elif "category" in parentTagName:
        sectionName = "{} (id={})".format(getReadableName(parentTagName), parentTag.get("id"))
    elif parentTagName == "enum-value":
        sectionName = childChar + "Field Option (value={})".format(parentTag.get("value"))
    elif parentTagName.endswith("-sect"):
        index = parentTag.get("index")
        if parentTagName == "fm-sect":
            parentTagName = parentTag.parent.name

        sectionNameFromTag = parentTagName[:parentTagName.find('-')]
        sectionName = "Form Section: {} (index={})".format(
            sectionNameFromTag.capitalize(), index)
        if parentTagName == "objective-sect":
            objPlanIdTag = parentTag.find("obj-sect-plan-id")
            if objPlanIdTag != None:
                sectionName = "Form Section: {} (plan-id={})(index={})".format(
                    sectionNameFromTag.capitalize(), objPlanIdTag.string, index)
        elif parentTagName == "objcomp-summary-sect":
            xaxis = parentTag.find("x-axis")
            yaxis = parentTag.find("y-axis")
            if xaxis != None and yaxis != None:
                x = xaxis.string
                y = yaxis.string
                sectionName = "Form Section: {}(x) vs {}(y) Summary (index={})".format(
                    x.capitalize(), y.capitalize(), index)
            else:
                sectionName = "Form Section: Objective vs Competency Summary (index={})".format(
                    index)
        elif parentTagName == "perfpot-summary-sect":
            sectionName = "Form Section: Performance-Potential Summary (index={})".format(
                    index)

    elif parentTagName == "fm-competency":
        compIdTag = parentTag.find('fm-comp-id')
        if compIdTag != None:
            sectionName = childChar +"Competency (id={})".format(compIdTag.string)
    elif parentTagName == "fm-sect-config":
        sectionName = childChar + "Section Configuration"
    elif parentTagName == "scale-map-value":
        sectionName = "Scale Adjusted Calculation Mapping"
    else:
        sectionName = getReadableName(parentTagName)

    return sectionName

def getReadableName(tagName, includeTagName=False):
    readableName = ""
    words = tagName.split('-')
    for word in words:
        if word == 'sect': word = 'Section'
        elif word == 'intro': word = 'Introduction'
        elif word == 'comp' : word = 'Competency'
        elif word == 'fm' : word = " "
        readableName += word.capitalize() + " "
    if includeTagName:
        readableName = readableName.strip() + " (" + tagName +")" 
    else:
        readableName = readableName.strip()
        
    return readableName

def deriveSectionTagName(sectionName):
    sectionTagName = ""
    if sectionName.find("Form Section:") != -1:
        sectionName = sectionName[sectionName.find(":")+1:]
        sectionName = sectionName.lower().strip()+"-sect"
    elif sectionName.find(childChar) != -1:
        sectionName = sectionName[sectionName.find(childChar)+1:]
    
    sectionName = sectionName.strip()
    
    if sectionName == "Field Option":
        sectionTagName = "enum-value"
    elif sectionName == "Competency":
        sectionTagName = "fm-competency"
    elif sectionName == "Form Section: Performance-Potential Summary":
        sectionTagName = "perfpot-summary-sect"
    elif sectionName.find("(x) vs ") > -1:
        sectionTagName = "objcomp-summary-sect"
    elif sectionName == "Scale Adjusted Calculation Mapping":
        sectionTagName = "scale-map-value"
    elif sectionName == "Section Configuration":
        sectionTagName = "fm-sect-config"
    else:
        words = sectionName.casefold().split(' ')
        sectionTagName = "-".join(words)
    
    return sectionTagName

def getMissingLangs(labelTag, allLangIds):
    missingLangs = []
    parentTag = labelTag.parent
    for lang in allLangIds:
        if parentTag.find(attrs={"xml:lang" : lang}, recursive=False) == None:
            missingLangs.append(lang)
    return missingLangs                

def getLangTagOf(tag, lang, langTagName = "label"):
    langTag = None
    for childTag in tag.children:
        childTagName = childTag.name
        if childTagName == langTagName and childTag.has_attr("xml:lang") and childTag.get("xml:lang") == lang:
            langTag = childTag

    return langTag

def getDefaultTitle(tag, en_US=False, labelOnParent = False):
    tagLabel = ""
    tagLabelDef = ""
    tagLabelEng = ""
    tagName = ""
    childrenTags = []
    if labelOnParent:
        childrenTags = tag.parent.children
        tagName = tag.name
    else:
        childrenTags = tag.children
        

    childTagName = None
    for childTag in childrenTags:
        childTagName = childTag.name
        if isinstance(childTag, NavigableString):
            continue
        if (tagName!= "" and childTagName == tagName) or tagName == "":
            if (not childTag.has_attr("xml:lang") 
                and not childTag.has_attr("lang") 
                and not childTag.has_attr("id")
                and not childTag.has_attr("rule")):
                tagLabelDef = childTag.string
            
            if (childTag.get("xml:lang") == "en-US" 
                or childTag.get("lang") == "en_US"):
                tagLabelEng= childTag.string
            
            if not en_US:
                if tagLabelDef != None and tagLabelDef != "":
                    tagLabel = tagLabelDef
                    break
                elif tagLabelEng != None and tagLabelEng != "":
                    tagLabel = tagLabelEng
                    break
            else:
                if tagLabelEng != None and tagLabelEng != "":
                    tagLabel = tagLabelEng
                    break
                elif tagLabelDef != None and tagLabelDef != "":
                    tagLabel = tagLabelDef
                    break

            # If label in default lang is empty or missing, take label from default locale and put for default tag
            if tagLabel == "" or tagLabel == None:
                if childTag.get("xml:lang") == sysDefaultLang  or childTag.get("lang") == sysDefaultLang:
                    tagLabel= childTag.string
            
            if tagName == "mapto-desc":
                forScore = tag.parent.find("mapto-score")
                if forScore != None:
                    tagLabel = "{} (for score={})".format(tagLabel, forScore.string)

    if tagLabel == "" or tagLabel == None:
        if tag.get("id") != None:
            tagLabel = "{} ({})".format(tagName, tag.get("id"))
        elif tag.get("for") != None:
            tagLabel = "{} ({})".format(tagName, tag.get("for"))
    
    return tagLabel

def createSheetsPerLang(workbook, sheetName, langs, baseHeaders):
    worksheets = []
    headerCols = len(baseHeaders)
    headerCellId = "{}1".format(get_column_letter(headerCols+1)) 
    separator = "_"
    for langId in langs:
        worksheetName = "{} ({})".format(sheetName, langId)
        if langId.find("-") > -1:
            separator = "-"
        else:
            separator = "_"
        if not worksheetName in worksheets:
            worksheet = workbook.create_sheet(worksheetName)
            worksheets.append(worksheetName)
            worksheet.append(baseHeaders)
            if langId.startswith("bs{}".format(separator)):
                langId = "bs"
            if langId != "en{}DEBUG".format(separator):
                try:
                    worksheet[headerCellId]="Label/Name in "+babel.Locale.parse(langId,sep=separator).english_name
                except:
                    worksheet[headerCellId]="Label/Name in "+ langId
            else:
                worksheet[headerCellId]="Label/Name in SF Debug Language"

def changeCellsStyle(cell, newStyleName):
    if cell.font.b:
        isInBoldFont = True
    else:
        isInBoldFont = False
    if cell.row == 1 and newStyleName == "EditableCellStyle" and cell.value != None:
        cell.style = "LockedCellStyle"
    else:
        cell.style = newStyleName
    if isInBoldFont or cell.row == 1:
        cell.font = boldFont
    if cell.row == 1:
        cell.border = headerCellBorder

def prepareAndExportWorkbook(workbook):
    global excelFileName
    gradfill = GradientFill(degree=float(90), stop=("FFFFFF", "FED8B1"))
    gradHeaderFill = GradientFill(degree=float(90), stop=("FFFFFF", "E8E8E8"))
    gradEditFill = GradientFill(degree=float(90), stop=("FFFFFF", "C7F6C7"))
    
    if 'HeadersStyle' not in workbook.named_styles:
        headerStyle = NamedStyle(name="HeadersStyle")
        headerStyle.font = boldFont
        headerStyle.fill = gradHeaderFill
        headerStyle.border = border
        workbook.add_named_style(headerStyle)
    if 'LockedCellStyle' not in workbook.named_styles:
        lockedCellStyle = NamedStyle(name="LockedCellStyle")
        lockedCellStyle.font = Font(color='00003F')
        lockedCellStyle.fill = gradfill 
        lockedCellStyle.border = border
        workbook.add_named_style(lockedCellStyle)
    if 'EditableCellStyle' not in workbook.named_styles:
        editableCellStyle = NamedStyle(name="EditableCellStyle")
        editableCellStyle.fill = gradEditFill
        editableCellStyle.border = border
        editableCellStyle.protection = Protection(locked=False, hidden=False)
        workbook.add_named_style(editableCellStyle)
    # Protect Entire Workbook
    workbook.security.workbookPassword = "...ApTrans..."
    workbook.security.lockStructure = True
    # Set other properties
    workbook.properties.title = "SF Translations Workbook - By TREXIMA™"  
    workbook.properties.category = "TranslationsWorkbook"   
    workbook.properties.subject = "SF Translations Workbook"
    workbook.properties.creator = "TREXIMA™ by Sandeep Kumar (Deloitte)"

    for ws in workbook:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        for row in ws:
            # row is a tuple of cell objects
            if row[0].row == 10:
                break
            #print(row[0].row)
            for cell in row:
                if cell.row == 1:
                    cell.style = 'HeadersStyle'
                colWidth = ws.column_dimensions[cell.column_letter].width
                cellsNeededWidth = len(str(cell.value))
                if colWidth <= cellsNeededWidth+2:
                    ws.column_dimensions[cell.column_letter].width = min(35, cellsNeededWidth+2)

        ws.protection.password = '...ApTrans...'
        ws.protection.sheet = True
        protectedCols = ['A','B']
        freezeCellsRange = ws['C2']
        if ws.title == sheetNameForPL or ws.title.startswith("DataModel"):
            protectedCols.append('C')
            protectedCols.append('D')
            freezeCellsRange = ws['E2']
        elif not ws.title.startswith("ObjectDefinitions"):
            protectedCols.append('C')
            protectedCols.append('D')
            freezeCellsRange = ws['E2']

        ws.protection.sort = False
        ws.auto_filter.ref = "A1:D2"
        ws.protection.autoFilter = False
        ws.protection.formatColumns = False

        ws.freeze_panes = freezeCellsRange
        lockedColsCount = len(protectedCols)
        for col in ws.iter_cols(min_col=1, max_col=lockedColsCount, min_row=1):
            for cell in col:
                changeCellsStyle(cell, 'LockedCellStyle')
        
        for col in ws.iter_cols(min_col=lockedColsCount+1, max_col=ws.max_column+2, min_row=1):
            for cell in col:
                changeCellsStyle(cell, 'EditableCellStyle')

        if ws.title == sheetNameForPM:
            for cell in ws["F"]:
                changeCellsStyle(cell, 'LockedCellStyle')

    # Export File Name and Localtion
    timestampSuffix = time.strftime("%a_%d%b_%Y_%Hh%Mm%Ss", time.localtime())
    excelFileName = fileSaveDir + "\\SF_EC_Translations_"+timestampSuffix+".xlsx"
    
    excelFileName = gui.filesavebox(msg="Save Translations Export File As", title="Save As", default=excelFileName)
    workbook.save(excelFileName)
    progressMsg = "The generated Translations Workbook will open in a moment!"
    logTheProgress(100, progressMsg)
    os.startfile(excelFileName)
    progress.forget()

def appendAsHeaderRow(worksheet, cellValues):
    if worksheet != None:
        worksheet.append(cellValues)
        last_row = worksheet[worksheet.max_row]
        for cell in last_row:
            cell.style = 'BoldCellStyle'

def findInDataModelReferencesOf(tagName):
    references = []
    tagIds = []
    
    # Loop through all Data Model's XML content (called Soup)
    for soup in soups:
        parentTagId = ""
        parentTagName = ""
        matchingTags = soup.find_all(tagName)
        
        for matchingTag in matchingTags:
            tagId = matchingTag.get('id')
            if(tagName == 'trigger-rule'):
                tagId = matchingTag.get('rule')+ '('+matchingTag.get('event')+')'
            parentTag = matchingTag.parent
            parentTagId = parentTag.get('id')
            parentTagName = parentTag.name
            visibility = parentTag.get('visibility')
            if visibility == 'none' or parentTagName in tagsToBeIgnored:
                continue
            sectionName = ""
            fieldName = getDefaultTitle(parentTag)
            if parentTagName in employeeProfileTags:
                sectionName = "Employee Profile -> "
                epParentTag = parentTag.parent
                bgElementId = epParentTag.get('id')
                if epParentTag.name == "background-element":
                    sectionName += "Background Element -> " + getDefaultTitle(parentTag) + " ("+ bgElementId +")"
                elif parentTagName == "userinfo-element":
                    sectionName += "User-Info Element"

            elif parentTagName.startswith("hris"): # an EC field
                
                sectionName = "Employee Central (GLOBAL) -> "
                hrisParentTag = parentTag.parent
                
                if hrisParentTag.name == "hris-section":
                    ecSectionName = getDefaultTitle(hrisParentTag)
                    ecPortlet = hrisParentTag.parent # hris-element
                    ecPortletName = getDefaultTitle(ecPortlet)
                    if ecPortlet.name == "country":
                        sectionName = "Employee Central ("+hrisParentTag.parent.parent.get("id")+") -> "
                    sectionName += ecPortletName + " (Portlet) -> "
                    sectionName += ecSectionName + " (Section)"
                elif hrisParentTag.name == "hris-element":
                    if hrisParentTag.parent.name == "country":
                        sectionName = "Employee Central ({}) -> ".format(hrisParentTag.parent.get("id"))
                    sectionName += getDefaultTitle(hrisParentTag) + " (Portlet)"

            reference = "{} -> {} | ".format(sectionName, fieldName) 
            tagIds.append(tagId)
            references.append(reference)
    return {"ids":tagIds, "references":references}

def findColumnLetterForLang(worksheet, lang):
    column_letter = None
    for cell in worksheet["1:1"]:
        if lang in cell.value:
            cell.column_letter
    return column_letter

def executeExport():
    global tagsToBeIgnored
    global exportExecuteBtn
    global picklistIds
    global picklistReferences
    global sysDefaultLang

    progressVal = 0
    #countdown(1)
    if excelFileName != None:
        openExisting = gui.ynbox("An Export is already available from a recent execution. Do you want to open it? Click 'No' to start afresh!",
                  "Recent Export Already Exists", ('Yes, Open Recent Export', 'No, Execute New Export'))
        if openExisting:
            os.startfile(excelFileName)
            return
        else:
            logTheProgress(progressVal, "Ready? Choose your desired options for this export when prompted!")
            #TODO - Remove below settings - 2 lines after testing
            progress['maximum'] = 100
            progress.pack(pady = 250)

    # Excel Workbook Initialization for Export
    #headerFont = Font(underline="single")
    workbook = Workbook()
    workbook.iso_dates = True
    boldFont = Font(color='00000000', bold=True, size=12)
    if 'BoldCellStyle' not in workbook.named_styles:
        boldCellStyle = NamedStyle(name="BoldCellStyle")
        boldCellStyle.font = boldFont
        workbook.add_named_style(boldCellStyle)

    # Finding references for all picklists in the data models and in MDF Object Definitions
    
    
    exportPicklistsTranslations = gui.ynbox("Do you want to export translations for all Picklists?",
        " Export Picklists' Translations?")
    
    if exportPicklistsTranslations:
        foundResult = findInDataModelReferencesOf("picklist")
        picklistIds = foundResult.get("ids")
        picklistReferences = foundResult.get("references")
        progressVal += len(soups)
        logTheProgress(progressVal, "Picklist references collected from all data models!")
        picklistFromCSV = gui.ynbox("Do you want to export all Picklists from Picklist-Values CSV export?. Choose yes, if you have exported the Picklist from SF. "+
                                    "Clicking No will export the picklists from API", " Export Picklists from latest picklist values export?")
        

    # Find list of active locales (language packs) in the SF instance
    defaultLangs = ["defaultValue", "localized", "en_DEBUG"]
    checklist = ["label_{}".format(lang) for lang in defaultLangs]
    properties = []
    if sfOdataService != None:
        properties = sfOdataService.schema.entity_type('PickListValueV2').proprties()
    activeLocales = []
    for prop in properties:
        propName = prop.name
        if propName.startswith('label_') and not (propName in checklist):
            activeLocales.append(propName[6:])
    
    labelKeysLangs = []
    for key in labelKeysFileHeaders:
        if key not in ['label_key', 'default']:
            labelKeysLangs.append(key)
    if (len(labelKeysLangs) > 0 and 
        len(set(activeLocales).intersection(labelKeysLangs)) != len(activeLocales)):
        activeLocales = labelKeysLangs

    if len(activeLocales) == 0:
        activeLocales = defaultLangs

    localesForExport = gui.multchoicebox("Choose Locales for the export of the translations in those locales",
                            "Choose Locales", activeLocales,[n for n in range(len(activeLocales))])
    if localesForExport == None or len(localesForExport) == 0:
        localesForExport = activeLocales
    
    
    indexOfDefLang = 0
    try:
        indexOfDefLang = localesForExport.index(sysDefaultLang)
    except:
        localesForExport.insert(0, sysDefaultLang)
        indexOfDefLang = 0
    
    sysDefaultLang = gui.choicebox("Choose SF instance's Default Language","Default Language",localesForExport,indexOfDefLang)
    
    if sysDefaultLang in localesForExport:
        localesForExport.remove(sysDefaultLang)
        localesForExport.insert(0,sysDefaultLang)

    allEntities = []
    if sfOdataService != None:
        allEntities = [es.name for es in sfOdataService.schema.entity_sets]
    # Proceed with extraction of translations for core MDF objects
    exportMDFObjDefns = gui.ynbox("Do you want to export translations from MDF Objects' Definitions?"," Export Translations from MDF Objects' Definitions?")
    
    if exportMDFObjDefns:
        progressVal += 2
        logTheProgress(progressVal, 
            "Extraction of Translations for Labels of MDF Objects' and Fields' starts...")

        mdfObjEntities = ["Position", "PaymentInformationDetailV3", "FOCompany",
                    "FOBusinessUnit","FODivision", "FODepartment", "FOJobCode",
                    "FOJobFunction","FOCostCenter","FOPayGroup"]
        
        customObjs = [entity for entity in allEntities if entity.startswith("cust_")]
        mdfObjEntities.extend(customObjs)

        progressIncr = 0
        mdfObjEntities = gui.multchoicebox("Choose Objects for Field Labels Translations Export", "Choose Objects", mdfObjEntities, preselect=[i for i in range(len(mdfObjEntities))])
        if mdfObjEntities == None: 
            mdfObjEntities = []
        else:
            progressIncr = 5/len(mdfObjEntities)
            baseHeaders = ["Object","Id"]
            createSheetsPerLang(workbook,"ObjectDefinitions", localesForExport, baseHeaders)
        
        for objectName in mdfObjEntities:
            progressVal += progressIncr
            langParam = ""
            objId = objectName
            if objectName == "FOCompany":
                objId = "LegalEntity"
            logTheProgress(progressVal,
                "Extracting of Translations for the MDF Object - {}".format(objId))
            
            for lang in localesForExport:
                ws = workbook["ObjectDefinitions ({})".format(lang)]
                langParam = "?sap-language="+lang
                metadataXML = sfOdataService.http_get("/{}/$metadata{}".format(objectName, langParam)).text
                msoup = BeautifulSoup(metadataXML, "xml")
                entitySets = msoup.find_all("EntitySet")
                entitySetTag = entitySets[0]
                ws.append([entitySetTag.get("Name"), objId, entitySetTag.get("sap:label")])
                last_row = ws[ws.max_row]
                for cell in last_row:
                    cell.style = 'BoldCellStyle'
                properties = msoup.find_all("Property")
                for prop in properties:
                    if prop["sap:visible"] == "true":# This seems to be true also for hidden fields!
                        ws.append([objId, prop["Name"], prop["sap:label"]])
                        if exportPicklistsTranslations:
                            picklistId = prop.get("sap:picklist")
                            if picklistId != None:
                                picklistIds.append(picklistId)
                                picklistReferences.append("MDF Object ({}) -> Field ({})".format(objId, prop["Name"]))

    x = 0
    progressIncr = 0

    if exportPicklistsTranslations and not picklistFromCSV:
        # Extract Picklists using SF OData API - Ensure that the ongoing PL migration is handled
        logTheProgress(progressVal+1, "Extraction of Picklists with Translations is in progress...")

        batchSize = 10 # No. of Picklists to fetch in each API call
        picklistItems = []
        
        migratedLegacyPLs = sfOdataService.entity_sets.PickListV2.get_entities().count().filter("legacyPickListId ne null").execute()
        totalLegacyPLs = 0
        if migratedLegacyPLs == 0:
            # Extract all the Legacy Picklists
            totalLegacyPLs = sfOdataService.entity_sets.Picklist.get_entities().count().execute()
            progressIncr = 10*batchSize/totalLegacyPLs
            while x < totalLegacyPLs:
                progressVal += progressIncr
                progress['value'] = progressVal
                appUI.update()
                picklistItems.extend(sfOdataService.entity_sets.Picklist.get_entities().top(batchSize).skip(x).expand('picklistOptions/picklistLabels').execute())
                x = x + batchSize

        # Extract all the MDF Picklists (which would include migrated legacy picklists, if any)
        x = 0    
        totalMDFPLs = sfOdataService.entity_sets.PickListV2.get_entities().count().execute()
        progressIncr = 10*batchSize/totalMDFPLs
        while x < totalMDFPLs:
            progressVal += progressIncr
            progress['value'] = progressVal
            appUI.update()
            picklistItems.extend(sfOdataService.entity_sets.PickListV2.get_entities().top(batchSize).skip(x).expand('values').execute())
            x = x + batchSize
        
        print("Total PLs = {} out of which Total Legacy PLs = {} and Total MDF PLs = {}".format(len(picklistItems), totalLegacyPLs, totalMDFPLs))

        logTheProgress(progressVal + 1, "Picklists have been extracted. Adding to the export file...")

        plOptionsProp = "picklistOptions"
        plHeaders = ["Reference","Picklist Id","Option's Unique Code", "Option ID"]
        for locale in localesForExport:
            plHeaders.append("Option Label ({})".format(locale))
        ws = workbook.create_sheet(sheetNameForPL)
        ws.append(plHeaders)

        for picklistItem in picklistItems:
            isMDFPL = False
            try:
                picklistItem.__getattr__("values")
                picklistId = picklistItem.id
                plOptionsProp = "values"
                isMDFPL = True
            except AttributeError:
                picklistId = picklistItem.picklistId
                plOptionsProp = "picklistOptions"

            references = ""
            while picklistIds.count(picklistId) > 0:
                index = picklistIds.index(picklistId)
                picklistIds.pop(index)
                reference = picklistReferences.pop(index)
                references += reference
            if references == "":
                continue
            
            for picklistOption in picklistItem.__getattr__(plOptionsProp):
                externalCode = picklistOption.externalCode
                rowData = [references, picklistId, externalCode]
                labels = []
                locale = ""
                if isMDFPL:
                    optionId = picklistOption.optionId
                    for locale in localesForExport:
                        label = picklistOption.__getattr__("label_{}".format(locale))
                        if locale == sysDefaultLang:
                            labels.insert(0,label)
                        else:
                            labels.append(label)
                else:
                    optionId = picklistOption.id
                    for locale in localesForExport:
                        for picklistOptionLabel in picklistOption.picklistLabels:
                            localeOfPlOption = picklistOptionLabel.locale 
                            label = picklistOptionLabel.label
                            if locale == sysDefaultLang:
                                labels.insert(0,label)
                                break
                            elif locale == localeOfPlOption:
                                labels.append(label)
                                break
                rowData.append(optionId)
                rowData.extend(labels)
                ws.append(rowData)
        print("All Picklists with references::::::::::::::::::::::::")
        print(*picklistIds)
        
    elif exportPicklistsTranslations:
        # Add the exported Picklist CSV as sheet to the Workbook
        csvFiles = filedialog.askopenfilenames(initialdir=os.getcwd(), title="Open latest exported picklist CSV file", 
                                        filetypes=[("CSV File","*.csv")])
        valuesMatrix = getValuesAsMatrixFromCSV(csvFiles[0])
        ws = workbook.create_sheet(sheetNameForPL)
        ws.column_dimensions[get_column_letter(1)].width = 75
        columnNum = 0
        headerRow = valuesMatrix[0]

        for header in headerRow:
            columnNum += 1
            if columnNum > 1 and header not in ["id","values.externalCode"] and not header.startswith("values.label."):
                ws.column_dimensions[get_column_letter(columnNum)].hidden = True
        
        
        rowNum = 0
        picklistId = ""
        references = ""
        for row in valuesMatrix:
            rowNum += 1
            if picklistId != row[1]:
                references = ""
                picklistId = row[1]
                while picklistIds.count(picklistId) > 0:
                    index = picklistIds.index(picklistId)
                    picklistIds.pop(index)
                    reference = picklistReferences.pop(index)
                    if reference not in references:
                        references += reference
            
            if rowNum < 3:
                row[0] = "References in EC"
            else:
                row[0] = references
            ws.append(row)
            
            
            if rowNum > 2 and references == "":
                ws.row_dimensions[rowNum].hidden = True
        
        print("All Picklists with references::::::::::::::::::::::::")
        print(*picklistIds) 
   


    # Proceed with extracting translations for Foundation Data
    exportFOTranslations = True
    removeHTMLTags = False

    if isPMGMIncluded:
        removeHTMLTags = gui.ynbox("Do you want to remove html tags, if contained in any translation?","Remove HTML tags?")
        exportFOTranslations = gui.ynbox("Do you want to also export translations for Foundation Objects (EC Relevant)?"," Export Translations for Foundation Objects?")
        exportECModelTranslations = True
        if isSDMIncluded:
            exportECModelTranslations = gui.ynbox("Do you want to also export translations for EC relevant data model elements and fields?"," Export Translations for EC Sections?")
        
        if not exportECModelTranslations:
            tagsToBeIgnored = ["hris-section", "hris-element", "hris-field", "format-group", "format"]
    else:
        exportFOTranslations = gui.ynbox("Do you want to export translations for Foundation Objects (EC Relevant)?"," Export Translations for Foundation Objects?")

    if exportFOTranslations:
        
        progressVal = progressVal+3
        logTheProgress(progressVal, "Extracting translations for Foundation Objects...")
        
        objectsForSelection = ["AlertMessage", "CustomPayType"] 
        filteredEntities = [entity for entity in allEntities 
                        if (entity.startswith("FO") or entity.startswith("cust_")) and not entity.startswith("FOW") and not entity.endswith("DEFLT")]
        
        # Ensure the High Priority FO's stay at top of thelist
        filteredEntities.remove("FOEventReason")
        filteredEntities.remove("FOLocation")
        filteredEntities.remove("FOPayComponent")
        filteredEntities.remove("FOFrequency")

        filteredEntities.insert(0, "FOEventReason")
        filteredEntities.insert(1, "FOLocation")
        filteredEntities.insert(2, "FOPayComponent")
        filteredEntities.insert(3, "FOFrequency")

        objectsForSelection.extend(filteredEntities)
        objectsToBeExported = gui.multchoicebox("Choose object(s) you need translations to be exported for. Click to select/unselect.", 
            "Choose Foundation and Configuration Objects For Export", objectsForSelection, [n for n in range(len(objectsForSelection))])
        if objectsToBeExported == None:
            objectsToBeExported = []
        else:
            entitySetsDict = sfOdataService.entity_sets.__dict__['_entity_sets']
            ws = workbook.create_sheet("FOAndConfigData")
            systemLangs = localesForExport
            headers = ["FO/Config Name","Code", "Field", "Default Label"]
            headers += ["Label in {}".format(babel.Locale.parse(lang).english_name) for lang in systemLangs]
            ws.append(headers)

            for objName in objectsToBeExported:
                entityMetadata = sfOdataService.schema.entity_type(objName)
                nav_properties = entityMetadata.nav_proprties
                nav_properties = [prop.name for prop in nav_properties]
                key_property = entityMetadata.key_proprties[0].name
                all_properties = [proprty.name for proprty in entityMetadata.proprties()]
                translatableProps = []
                translatableFields = []
                for prop in all_properties:
                    for lang in systemLangs:
                        if prop.endswith("_"+lang):
                            translatableProps.append(prop)
                            fieldName = prop[:prop.find("_")]
                            if fieldName not in translatableFields:
                                translatableFields.append(fieldName)
                isLegacyFO = False
                objects = []
                try:
                    if len(translatableProps) > 0:
                        objects = entitySetsDict[objName].get_entities().execute()
                    elif "nameTranslationNav" in nav_properties:
                        isLegacyFO = True
                        objects = entitySetsDict[objName].get_entities().expand('nameTranslationNav,descriptionTranslationNav').execute()
                    
                    if objects != None and len(objects) > 0:
                        
                        for obj in objects:
                            rowCellValues = [objName]
                            key = ""
                            try:
                                key = obj.__getattr__(key_property)
                            except:
                                continue
                            
                            if isLegacyFO:
                                rowCellValues.append(key)
                                nameTransNav = obj.__getattr__('nameTranslationNav')
                                if nameTransNav != None:
                                    rowCellValues.append("name")
                                    rowCellValues.append(nameTransNav.__getattr__("value_defaultValue"))
                                    for lang in systemLangs:
                                        try:
                                            rowCellValues.append(nameTransNav.__getattr__("value_"+lang))
                                        except:
                                            continue
                                else:
                                    rowCellValues.append("name")
                                    rowCellValues.append(obj.__getattr__("name"))
                                ws.append(rowCellValues)

                                rowCellValues = [objName, key]
                                descTransNav = obj.__getattr__('descriptionTranslationNav')
                                if descTransNav != None:
                                    rowCellValues.append("description")
                                    rowCellValues.append(descTransNav.__getattr__("value_defaultValue"))
                                    for lang in systemLangs:
                                        try:
                                            rowCellValues.append(descTransNav.__getattr__("value_"+lang))
                                        except:
                                            continue
                                else:
                                    rowCellValues.append("description")
                                    rowCellValues.append(obj.__getattr__("description"))
                                ws.append(rowCellValues)
                            else:
                                for translatableField in translatableFields:
                                    rowCellValues = [objName, key, translatableField]
                                    try:
                                        defaultVal = obj.__getattr__(translatableField+"_defaultValue")
                                    except pyodata.exceptions.PyODataException as e:
                                        defaultVal = None
                                    if defaultVal != None:
                                        rowCellValues.append(defaultVal)
                                    else:
                                        rowCellValues.append(obj.__getattr__(translatableField))
                                    for lang in systemLangs:
                                        rowCellValues.append(obj.__getattr__(translatableField+"_"+lang))
                                    ws.append(rowCellValues)

                except pyodata.exceptions.HttpError as e:
                    print("Error During API Call for {} => {} ".format(objName, e.response.text))

    # Proceed with extracting translations from Data Models
    progressVal = progressVal+7
    logTheProgress(progressVal, 
        "Preparing multi-sheet excel file with translations for UI Elements of SF EC...")

    systemXMLLangs = [x.replace("_","-") for x in localesForExport]
    allLangIds = []
    worksheetName = None

    if askToChooseDMLangs:
        for soup in soups:
            langsPerDM = []
            langAttrTags = soup.find_all("label")
            for langAttrTag in langAttrTags:
                langId = langAttrTag.get('xml:lang')
                if langId != None and langsPerDM.count(langId) == 0:
                    langsPerDM.append(langId)
            selectedlangsIndices = [langsPerDM.index(y) for y in langsPerDM if y in systemXMLLangs]
            
            allLangIds = [x for x in langsPerDM if x not in allLangIds]
            
            if len(allLangIds) > 0:
                allLangIds = gui.multchoicebox(msg='Choose Languages for the Translations export.',
                                title='Choose Languages For Export', 
                                choices=langsPerDM, preselect=selectedlangsIndices)
            if allLangIds is None or len(allLangIds) == 0: 
                allLangIds = langsPerDM
            
            if len(allLangIds) != len(langsPerDM): 
                break
    else:
        allLangIds = [langId.replace("_","-") for langId in localesForExport]
    
    worksheets = []
    worksheet = None
    ws1 = None
    ws2 = None

    headers = ["Section","Element/Subsection","Field Id","Default Label"]
    if isPMGMIncluded:
        headers = ["Translation Type", "Template Name", "Section/Element/Subsection","Translatable Item/Field","Default Label","Label Key"]
        ws1 = workbook.create_sheet(sheetNameForPM)
        worksheets.append(sheetNameForPM)
        ws2 = workbook.create_sheet(sheetNameForGM)
        worksheets.append(sheetNameForGM)

    for langId in allLangIds:
        if isPMGMIncluded:
            lang = langId.replace('-','_')
            if langId.startswith("bs-"):
                    langId = "bs"
            if langId != "en-DEBUG":
                headers.append("Label in {} ({})".
                    format(babel.Locale.parse(langId,sep='-').english_name, lang))
            else:
                headers.append("Label SF Debug (en-DEBUG)")
    
    if isPMGMIncluded:
        ws1.append(headers)
        headers.remove("Label Key")
        ws2.append(headers)
        
    if isSDMIncluded:
        headers = ["Section","Element/Subsection","Field Id","Default Label"]
        for langId in allLangIds:
            worksheetName = "DataModel ({})".format(langId)
            if not worksheetName in worksheets:
                worksheet = workbook.create_sheet(worksheetName)
                worksheets.append(worksheetName)
                worksheet.append(headers)
                if langId.startswith("bs-"):
                    langId = "bs"
                if langId != "en-DEBUG":
                    worksheet["E1"]="Label in {}".format(babel.Locale.parse(langId,sep='-').english_name)
                else:
                    worksheet["E1"]="Label in SF Debug (en-DEBUG)"

    progressVal = progressVal+5
    progressMsg = "Sheets for various languages are created. Going to extract translations..."
    logTheProgress(progressVal, progressMsg)

    filesCount = len(soups)
    progressIncr = (90 - progressVal)/filesCount
    configName = ""
    highlightTags = ["succession-data-model", "background-element", "userinfo-element", "hris-element", "hris-section"]
    
    isFirstRowAdded = False
    for soup in soups:
        configName = getSFSpecificName(soup)
        standardDMSoup = soupsDict.get("{} {}".format(KEYWORD_SAP_STANDARD, configName))
        tagsWithTranslations = soup.find_all(translatableTags)
        progressVal += progressIncr
        progress['value'] = progressVal
        isPMGMSoup = False
        translationFeature = "Manage Templates"

        if soup.find("sf-pmreview") != None:
            for key in soupsDict:
                if soup == soupsDict[key]:
                    configName = key
             
            worksheetName = sheetNameForPM
            isPMGMSoup = True
            translationFeature = "Manage Templates -> Performance Review"
        elif soup.find("obj-plan-template") != None:
            worksheetName = sheetNameForGM
            isPMGMSoup = True
            if soup.find("obj-plan-type").string == "Development":
                translationFeature = "Manage Templates -> Development"
            else:
                 translationFeature = "Manage Templates -> Goal Plan"
        else:
            worksheetName = None
            isPMGMSoup = False
        
        if worksheetName in worksheets:
            worksheet = workbook[worksheetName]
        
        progressMsg = "Extracting translations from {}".format(configName)
        logTheProgress(progressVal+1, progressMsg)
        
        prevCountry = ""
        headerCreationMap = {}
        prevParentTag = None
        prevTagName = None
        addedCountries = []
        addedCountriesPerWS = {}

        for tagWithTranslations in tagsWithTranslations:
            parentTag = tagWithTranslations.parent
            parentTagName = parentTag.name
            subSectionName = parentTagName
            parentTagId = parentTag.get('id')
            visibility = parentTag.get('visibility')
            
            if visibility == 'none' or parentTagName in tagsToBeIgnored:
                #print("Tag Name and visibility".format(tagWithTranslations.name, visibility))
                continue  # Ignore this tag for the export
            
            matchingTagInStdDM = None
            if standardDMSoup != None:
                matchingTagInStdDM = standardDMSoup.find(name=parentTagName, attrs={'id':parentTagId})

            defaultLabel = getDefaultTitle(tagWithTranslations, False, True)
            grandParent = parentTag.parent
            sectionName = ""
            countryCode = ""
            skipThisCountry = False
            if parentTagName in employeeProfileTags:
                sectionName = "Employee Profile"
            elif parentTagName.startswith("hris") or parentTagName == "format":
                if configName.find(" CSF ") > -1:
                    isGrandChildOfCountry = False
                    if grandParent.name == "hris-section":
                        countryCode = grandParent.parent.parent.get('id') 
                    elif grandParent.name == "hris-element" or grandParent.name == "format-group":
                        countryCode = grandParent.parent.get('id')
                        isGrandChildOfCountry = True
                    elif parentTag.name == "hris-element" or parentTag.name == "format-group":
                        countryCode = grandParent.get('id')

                    if countryCode == "": 
                        print("SOMETHING WRONG: How Come this happened?")
                    if len(activeCountries) > 0 and countryCode not in activeCountries:
                        skipThisCountry = True
                    sectionName = "{} ({})".format(configName, countryCode)
                else:
                    sectionName = "{}".format(configName)
            else:
                sectionName = "{}".format(configName)
                subSectionName = getModuleSpecificName(parentTag)
                field = getReadableName(tagWithTranslations.name, True)
                #print("YES, this is where I should be form PMGM module - For Tag: '{}'".format(field))

            if isPMGMSoup:
                if parentTag == prevParentTag and prevTagName == tagWithTranslations.name:
                    continue
                if worksheet != None:
                    if (removeHTMLTags and
                            "<" in defaultLabel and "</" in defaultLabel):
                        soup = BeautifulSoup(defaultLabel, features="lxml")
                        defaultLabel = soup.get_text()
                    
                    if worksheetName == sheetNameForPM and not isFirstRowAdded:
                        rowValues = [translationFeature, sectionName, "General Settings","Form Name",sectionName,""]
                        worksheet.append(rowValues)
                        isFirstRowAdded = True
                    
                    rowValues = [translationFeature, sectionName, subSectionName, field, defaultLabel]
                    
                    for langId in allLangIds:
                        lang = langId.replace('-','_')
                        langLabelTag = parentTag.find(attrs={"lang" : lang}, recursive=False)
                        if langLabelTag == None:
                            msgKey = tagWithTranslations.get("msgKey")
                            if msgKey == None:
                                msgKey = tagWithTranslations.get("msgkey")
                            if msgKey == None:
                                langLabel = ""
                                msgKey = ""
                            else:
                                labelKeyRowDict = labelKeysDict.get(msgKey)
                                if labelKeyRowDict != None:
                                    langLabel = labelKeyRowDict.get(lang, "")
                                else:
                                    langLabel = ""
                                    msgKey = "{} - CONFIG ERROR (referred but missing in FormLabelKeys CSV)".format(msgKey)
                            if (worksheetName == sheetNameForPM and 
                                msgKey != "" and 
                                msgKey not in rowValues): rowValues.append(msgKey)
                        else:
                            langLabel = langLabelTag.string
                        
                        if (removeHTMLTags and
                            "<" in langLabel and "</" in langLabel):
                            soup = BeautifulSoup(langLabel, features="lxml")
                            langLabel = soup.get_text()
                        
                        rowValues.append(langLabel)
                    worksheet.append(rowValues)
                
                prevParentTag = parentTag
                prevTagName = tagWithTranslations.name
            elif not skipThisCountry:
                
                if parentTag != prevParentTag:
                    missingLangs = getMissingLangs(tagWithTranslations, allLangIds)
                    for missingLang in missingLangs:
                        worksheetName = "DataModel ({})".format(missingLang)

                        if worksheetName in worksheets:
                            worksheet = workbook[worksheetName]
                            if ((parentTagName.startswith("hris") or parentTagName == "format") and
                            configName.find(" CSF ") > -1 and isGrandChildOfCountry):
                                countriesAlreadyAdded = addedCountriesPerWS.get(worksheetName)
                                if countriesAlreadyAdded==None or countryCode not in countriesAlreadyAdded:
                                    appendAsHeaderRow(worksheet, [sectionName, "country", countryCode, "", ""])
                                    addedCountries.append(countryCode)
                                    addedCountriesPerWS[worksheetName] = addedCountries
                            standardLabel = ""
                            
                            if matchingTagInStdDM != None:
                                langTag = getLangTagOf(matchingTagInStdDM, missingLang)
                                if langTag != None:
                                    standardLabel = langTag.string

                            rowValues = [sectionName, subSectionName, parentTagId, defaultLabel, standardLabel]
                            if parentTagName in highlightTags:
                                appendAsHeaderRow(worksheet, rowValues)
                            else:
                                worksheet.append(rowValues)
                prevParentTag = parentTag
                
                langAttr = tagWithTranslations.get('xml:lang')
                    
                if langAttr != None and langAttr != "":
                    worksheetName = "DataModel ({})".format(langAttr)
                    
                    if worksheetName in worksheets:
                        worksheet = workbook[worksheetName]
                        label = tagWithTranslations.string
                        
                        if ((parentTagName.startswith("hris") or parentTagName == "format") and
                            configName.find(" CSF ") > -1):  
                            if headerCreationMap.get(langAttr+countryCode) != grandParent.get('id') and grandParent.contents[1].name != "label":
                                if isGrandChildOfCountry and headerCreationMap.get(langAttr) != countryCode:
                                    headerCreationMap[langAttr] = countryCode
                                    appendAsHeaderRow(worksheet, [sectionName, "country", countryCode, "", ""])
                                if isGrandChildOfCountry or grandParent.name == "hris-section":
                                    headerCreationMap[langAttr+countryCode] = grandParent.get('id')
                                    appendAsHeaderRow(worksheet, [sectionName, grandParent.name, grandParent.get("id"), "", ""]) 
                            elif countryCode != prevCountry:
                                #print("Reset the Dictionary "+headerCreationMap[langAttr+countryCode])
                                headerCreationMap[langAttr+countryCode] = None
                            prevCountry = countryCode
                        
                        if worksheet != None:
                            rowValues = [sectionName, subSectionName, parentTagId, defaultLabel, label]
                            if parentTagName in highlightTags:
                                appendAsHeaderRow(worksheet, rowValues)
                            else:
                                worksheet.append(rowValues)

    progressMsg = "Translations has been extracted! Preparing the export file..."
    logTheProgress(progressVal+2, progressMsg)

    exportExecuteBtn["state"] = "disabled"
    exportExecuteBtn["image"] = doneIcon
    exportExecuteBtn["compound"] = tk.LEFT

    del workbook["Sheet"]
    prepareAndExportWorkbook(workbook)

def get_key(dict, val): 
    for key, value in dict.items():
         if val == value: 
             return key 
    return "key doesn't exist"

def on_click(button):
    button.configure(background="green")

def updatePicklistsLabels():
    picklists_sheet =translationsWb[sheetNameForPL]   

def executeImport():
    global importExecuteBtn
    global progressLog

    timestamp = time.strftime("%a_%d%b_%Y_%Hh%Mm%Ss", time.localtime())
    importLogs = "{}: Started Importing Translations\n\tFrom: {}\n\tTo Data Model Files:\n\t{}".format(timestamp, importFilePath, xmlFiles)
    worksheetNames = translationsWb.sheetnames
    modifiedSoups = []
    progressVal = 0
    progressIncr = 55/len(worksheetNames)
    gradHeaderFill = GradientFill(degree=float(90), stop=("FFFFFF", "E8E8E8"))

    if 'HeadersStyle' not in translationsWb.named_styles:
        headerStyle = NamedStyle(name="HeadersStyle")
        headerStyle.font = boldFont
        headerStyle.fill = gradHeaderFill
        headerStyle.border = border
        translationsWb.add_named_style(headerStyle)

    worksheetsToProcess = gui.multchoicebox("Choose Worksheets to process",
                            "Choose Worksheets", worksheetNames,[n for n in range(len(worksheetNames))])
    
    for wsName in worksheetsToProcess:
        progressVal += progressIncr
        progress['value'] = progressVal
        appUI.update_idletasks()
        progressMsg = ("Processing '{}' sheet from the Workbook to "+
            "update respective configuration file").format(wsName)
        logTheProgress(progressVal, progressMsg)
        ws = translationsWb[wsName]
        ws.protection.password = "...ApTrans..."
        ws.protection.disable()
        rowNum = 1
        colNum = 1
        maxColNum = ws.max_column
        
        if maxColNum < 2:
            print("Skipping invalid worksheet - "+wsName)
            continue # skip further processing of this custom WS which is not part of original WB generated by TREXIMA and move to next WS
        
        for cell in ws[1:1]:
            if cell.value == None or cell.value=="":
                maxColNum = colNum
                break
            colNum += 1
        changeLogHeaderCell = ws.cell(row=1, column=maxColNum)
        changeLogHeaderCell.value = "Change Log Identified During Import"
        changeLogHeaderCell.style = 'HeadersStyle'
        ws.column_dimensions[changeLogHeaderCell.column_letter].width = 75
        if wsName.startswith("DataModel"):
            langId = wsName[11:len(wsName)-1]
            #print(langId)
            rowNum = 1
            parentTag = None
            grandParentTag = None
            matchingTag = None
            while rowNum < ws.max_row:
                appUI.update_idletasks()
                rowNum += 1
                col1Cell = ws.cell(row=rowNum, column=1)
                col2Cell = ws.cell(row=rowNum, column=2)
                col3Cell = ws.cell(row=rowNum, column=3)
                dmRefCell = col1Cell.value
                isCSF = False
                if dmRefCell == None or dmRefCell == "":
                    break
                elif dmRefCell.find("(") > -1:
                    dmName = dmRefCell[:dmRefCell.find("(")-1]
                    dmRefCell = dmName.strip()
                    isCSF = True

                translatableItem = col2Cell.value
                tagId = col3Cell.value
                langLabelFromWB = ws.cell(row=rowNum, column=5).value
                if dmRefCell == "Employee Profile":
                    dmRefCell = "SFEC Succession Data Model"
                soup = soupsDict.get(dmRefCell)
                if soup != None:
                    if col1Cell.font.b and col2Cell.font.b and col3Cell.font.b:
                        if translatableItem == "country" or grandParentTag==None:
                            grandParentTag = soup
                        elif grandParentTag!= None :
                            grandParentTag = parentTag
                        parentTag = grandParentTag.find(name=translatableItem, attrs={'id':tagId, 'visibility':'both'})
                        if parentTag == None:
                            parentTag = soup.find(name=translatableItem, attrs={'id':tagId})
                        
                    
                    if parentTag == None:
                        parentTag = soup
                    
                    matchingTag = parentTag.find(name=translatableItem, attrs={'id':tagId, 'visibility':'both'})
                    if matchingTag == None:
                        matchingTag = parentTag.find(name=translatableItem, attrs={'id':tagId})
                        if matchingTag == None:
                            matchingTag = soup.find(name=translatableItem, attrs={'id':tagId})
                
                    if matchingTag == None:
                        #print("No matching tag found in the "+dmRefCell)
                        importLogs +="\n\nNo matching tag found in the {}".format(dmRefCell)
                        #print("Row {} in {} => Data Model -> TagName, tagId = {} -> {}, {}".format(rowNum, wsName, dmRefCell, translatableItem, tagId))
                        importLogs +="\n\nRow {} in {} => Data Model -> TagName, tagId = {} -> {}, {}".format(rowNum, wsName, dmRefCell, translatableItem, tagId)
                        ws.cell(row=rowNum, column=6).value = "No matching tag found in the "+dmRefCell
                        continue

                    labelTagName = "label"
                    matchingLabelTag = matchingTag.find(name=labelTagName, attrs={'xml:lang':langId})
                    
                    if matchingTag.find(labelTagName) == None:
                        # when "label" tags are missing for the matched tag, look for "instruction" tags
                        labelTagName = "instruction"
                        matchingLabelTag = matchingTag.find(name=labelTagName, attrs={'xml:lang':langId})
                        if matchingTag.find(name=labelTagName) == None:
                            matchingLabelTag = None
                            #print("There not any matching label or instruction tag found!")
                            #print("Row {} in {} => Data Model -> TagName, tagId = {} -> {}, {}".format(rowNum, wsName, dmRefCell, translatableItem, tagId))
                            labelTagName = "label" # To be created in next step
                    
                    if matchingLabelTag == None:
                        #print("Row {} in {} => Data Model -> TagName, tagId = {} -> {}, {}".format(rowNum, wsName, dmRefCell, tagName, tagId))
                        #print("New tag '{}' created".format(labelTagName))
                        # Create a new "label" tag for that
                        newLabelTag = soup.new_tag(labelTagName)
                        newLabelTag["xml:lang"] = langId
                        if langLabelFromWB != None:
                            newLabelTag.string = langLabelFromWB
                            matchingTag.insert(2, newLabelTag) # To be verified
                            if not soup in modifiedSoups:
                                modifiedSoups.append(soup)
                            importLogs +="\n\n{}: Row {} in Sheet '{}' -> \n'{}' language translation for the tag '{}' ADDED with ('{}' tag's) value as '{}'".format(timestamp, rowNum, wsName, langId, translatableItem, labelTagName, langLabelFromWB)
                            ws.cell(row=rowNum, column=6).value = "Translation Added to the DM with ('{}' tag's) value as '{}'".format(labelTagName, langLabelFromWB)
                    else:
                        oldLabel = matchingLabelTag.string
                        if langLabelFromWB != oldLabel:
                            if not soup in modifiedSoups:
                                modifiedSoups.append(soup)
                            if langLabelFromWB != None:
                                matchingLabelTag.string = langLabelFromWB
                                timestamp = time.strftime("%a_%d%b_%Y_%Hh%Mm%Ss", time.localtime())
                                importLogs +="\n\n{}: Row {} in Sheet '{}' -> \n'{}' language translation ('{}' tag's value) for the tag {} CHANGED from {} to {}.".format(timestamp, rowNum, wsName, langId, labelTagName, translatableItem,  oldLabel, langLabelFromWB)
                                ws.cell(row=rowNum, column=6).value = "Translation Changed in the DM from {} to {}".format(oldLabel, langLabelFromWB)
                else:
                    print("No soup found for {}".format(dmRefCell))

        elif wsName in [sheetNameForPM, sheetNameForGM]:
            parentTagName = None
            if wsName == sheetNameForPM:
                parentTagName = "sf-form"
            else:
                parentTagName = "obj-plan-template"
            
            
            labelKeyRowDicts = []
            defaultLang = None
            for labelKey in labelKeysDict:
                if defaultLang == None:
                    defaultLang = labelKeysDict[labelKey]['default']
                labelKeyRowDicts.append(labelKeysDict[labelKey])
            
            colNum = 1
            langKeysFromXL = []
            langStartColNum = 0
            while colNum < ws.max_column:
                langColHeader = ws.cell(row=1, column=colNum).value
                if "(" in langColHeader:
                    if langStartColNum == 0:
                        langStartColNum = colNum
                    langColKey = langColHeader[langColHeader.rfind("(")+1:langColHeader.rfind(")")]
                    langKeysFromXL.append(langColKey)
                colNum += 1
            rowNum = 1
            maxRowNum = 1
            
            parentSectionTag = None
            parentTag = None
            templateName = None
            attributeName = ""
            attributeValue = ""
            unmatchedRows = []
            unmatchedTemplates = []
            while rowNum < ws.max_row:
                appUI.update_idletasks()
                rowNum += 1
                templateName = ws.cell(row=rowNum, column=2).value
                if templateName == None:
                    maxRowNum = rowNum -1
                    break
                sectionName = ws.cell(row=rowNum, column=3).value
                translatableItem = ws.cell(row=rowNum, column=4).value
                changeColText = None
                soup = soupsDict.get(templateName)
                if soup != None:
                    if sectionName.rfind("(") != -1:
                        attributeName = sectionName[sectionName.rfind("(")+1:sectionName.rfind("=")]
                        attributeValue = sectionName[sectionName.rfind("=")+1:sectionName.rfind(")")]
                        sectionName = sectionName[:sectionName.find("(")]
                        sectionTagName = deriveSectionTagName(sectionName)
                        parentSectionTag = soup.find(sectionTagName, attrs={'{}'.format(attributeName):attributeValue})
                        if parentSectionTag == None:
                            parentSectionTag = soup.find("fm-sect", attrs={'{}'.format(attributeName):attributeValue})
                        parentTag = parentSectionTag
                    elif sectionName ==  childChar + "Section Configuration":
                        parentTag = parentSectionTag.find("fm-sect-config")
                    else:
                        parentTag = soup.find(parentTagName)
                    
                    if translatableItem.rfind("(") != -1:
                        translatableItem = translatableItem[translatableItem.rfind("(")+1:translatableItem.rfind(")")]
                    
                    if parentTag != None:
                        tag = parentTag.find(translatableItem)
                    else:
                        print("ISSSSUEEEEEEE:: Row {}: sectionName = {}, parentTag = {}, parentSectionTag = {}".format(
                        rowNum, sectionName, parentTag, parentSectionTag
                        ))

                    colNum = 1
                    langLabelDictFromXL = {}
                    for n in range(langStartColNum, maxColNum):
                        langColVal = ws.cell(row=rowNum, column=n).value
                        langKey = langKeysFromXL[n-langStartColNum]
                        langLabelDictFromXL[langKey] = langColVal
                    
                    if tag != None:
                        defLabelFromXL = ws.cell(row=rowNum, column=5).value
                        msgKeyFromXL = ws.cell(row=rowNum, column=6).value
                        tagName = tag.name
                        modifiedLangs = []
                        oldLabels = []
                        newLabels = []
                        if wsName == sheetNameForPM:
                            labelKeyDict = None 
                            if tag.has_attr("msgKey") or tag.has_attr("msgkey"):
                                lowercase = False
                                msgKey = tag.get("msgKey")
                                if msgKey == None:
                                    msgKey = tag.get("msgkey")
                                    lowercase = True
                                labelKeyDict = labelKeysDict.get(msgKey)
                                if labelKeyDict == None:
                                    importLogs += ("\n\nTag {} has msgKey {} but this msgKey is "+
                                        "missing in the FormLabelKeys file. Therefore, this msgKey "+
                                        "would be added to the FormLabelKeys file!").format(
                                       tagName, msgKey)
                                    if lowercase:
                                        del tag["msgkey"]
                                elif lowercase:
                                    del tag["msgkey"]
                                    tag["msgKey"] = msgKey
                                    if not soup in modifiedSoups:
                                        modifiedSoups.append(soup)

                            else:
                                sectionPart = sectionTagName.replace('fm-','').title().replace('-','_')
                                itemPart = tagName.replace('fm-sect-','').title()
                                #itemPart = itemPart[itemPart.rfind('-')+1:]
                                itemPart = itemPart.replace('-','')
                                msgKey = sectionPart+'_'+attributeValue+'_'+itemPart

                            if labelKeyDict != None:
                                for langKey in langKeysFromXL:
                                    labelFromCSV = labelKeyDict.get(langKey)
                                    labelFromXL = langLabelDictFromXL.get(langKey) 
                                    if labelFromCSV!= '' and labelFromCSV != labelFromXL:
                                        for labelKeyRowDict in labelKeyRowDicts:
                                            if labelKeyRowDict['label_key'] == msgKey:
                                                labelKeyRowDict[langKey] = labelFromXL
                                                modifiedLangs.append(langKey)
                                                oldLabels.append(labelFromCSV)
                                                newLabels.append(labelFromXL)
                                                break
                                if len(modifiedLangs) != 0:
                                    importLogs +="""\n\n{}: Row {} in Sheet '{}' -> Labels for '{}' languages changed in 
                                        FormLabelKeys CSV file for msgKey '{}' from '{}' to '{}'""".format(
                                        timestamp, rowNum, wsName, modifiedLangs, msgKey, oldLabels, newLabels)
                                    changeColText = """Translation changed in the FormLabelKeys CSV file for languages 
                                        '{}' to values '{}'.""".format(modifiedLangs, newLabels)       
                            else:
                                labelKeyRowDict = {}
                                langValues = [x for x in langLabelDictFromXL.values() if x!=None]
                                if len(langValues) > 0:
                                    for header in labelKeysFileHeaders :
                                        if header == 'label_key':
                                            labelKeyRowDict[header] = msgKey
                                            tag["msgKey"] = msgKey
                                            if not soup in modifiedSoups:
                                                modifiedSoups.append(soup)
                                        elif header == 'default':
                                            labelKeyRowDict[header] = defaultLang
                                        else:
                                            labelFromXL = langLabelDictFromXL.get(header)
                                            labelKeyRowDict[header] = labelFromXL
                                            modifiedLangs.append(header)
                                            #oldLabels.append(labelFromCSV)
                                            newLabels.append(labelFromXL)

                                    if len(modifiedLangs) != 0:
                                        importLogs +="""\n\n{}: Row {} in Sheet '{}' -> A new LabelKey '{}' was ADDED 
                                            to FormLabelKeys File with values '{}'""".format(
                                            timestamp, rowNum, wsName, msgKey, labelKeyRowDict)
                                        changeColText = """A new Label Key '{}' was ADDED to the FormLabelKeys CSV file 
                                            with values '{}'.""".format(msgKey, labelKeyRowDict)
                                        labelKeyRowDicts.append(labelKeyRowDict)
                        else:
                            if not tag.has_attr("lang"):
                                defLangValFromXML = tag.string
                                if defLabelFromXL == None: 
                                    defLabelFromXL = ""

                                if defLabelFromXL != "" and defLangValFromXML != defLabelFromXL:
                                    tag.string = "<![CDATA["+defLabelFromXL+"]]>"
                                    modifiedLangs.append("Default")
                                    oldLabels.append(defLangValFromXML)
                                    newLabels.append(defLabelFromXL)
                            
                            nextSiblingTags = tag.findNextSiblings(tagName)
                            matchingTag = tag
                            for langKey in langLabelDictFromXL:
                                langValFromXL = langLabelDictFromXL.get(langKey)
                                if langValFromXL == None or langValFromXL.strip() == "": 
                                    continue # If no input label provided in WB, then skip that cell
                                langKeyTagFound = False
                                for langAttrTag in nextSiblingTags:
                                    langTagKey = langAttrTag.get("lang")
                                    if langTagKey == langKey:
                                        matchingTag = langAttrTag
                                        langKeyTagFound = True
                                        langValFromXML = langAttrTag.string
                                        if  langValFromXML != langValFromXL:
                                            langAttrTag.string = "<![CDATA["+langValFromXL+"]]>"
                                            modifiedLangs.append(langKey)
                                            oldLabels.append(langValFromXML)
                                            newLabels.append(langValFromXL)
                                        break
                                if not langKeyTagFound:
                                    newLangKeyTag = soup.new_tag(tagName)
                                    newLangKeyTag["lang"] = langKey
                                    newLangKeyTag.string = "<![CDATA["+langValFromXL+"]]>"
                                    matchingTag.insert_after(newLangKeyTag)
                                    if not soup in modifiedSoups:
                                        modifiedSoups.append(soup)
                            
                            
                            if len(modifiedLangs) > 0:
                                changeColText = "Translation Changed for '{}' languages from '{}' to '{}'.".format(
                                modifiedLangs, oldLabels, newLabels)
                                importLogs +="""\n\n{}: Row {} in Sheet '{}' -> 
                                    \nTranslations for '{}' languages for the tag '{}' CHANGED from '{}' to '{}'""".format(
                                    timestamp, rowNum, wsName, modifiedLangs, tagName, oldLabels, newLabels)
                            
                                if not soup in modifiedSoups:
                                    #print(str(soup))
                                    modifiedSoups.append(soup)
                
                elif templateName != None:
                    unmatchedRows.append(rowNum)
                    if templateName not in unmatchedTemplates:
                        unmatchedTemplates.append(templateName)

                if changeColText != None:
                    ws.cell(row=rowNum, column=maxColNum).value = changeColText

            if len(unmatchedRows) > 0:
                retry = gui.ynbox(msg=("Template names {} mentioned in the rows {} (column B) of the "
                        "input translations workbook sheet '{}' does not match any of the input PMGM templates. "
                        "Please ensure that they match and then retry.").format(unmatchedTemplates, unmatchedRows, wsName), 
                        choices=["I will Edit & Retry", "Skip Unmatched Rows"])
                if retry:
                    resetExecutionProgress()
                    return
            
            if wsName == sheetNameForPM:
                # Write back to LabelKeys CSV file:
                newLabelKeysFile = fileSaveDir+"\\ReadyToImport_FormLabelKeys.csv"
                writer = csv.DictWriter(open(newLabelKeysFile, 'w', newline='', encoding="utf-8"), labelKeysFileHeaders)
                writer.writeheader()
                writer.writerows(labelKeyRowDicts)
                progressMsg = ("'Ready to Import' form label keys file "+
                    "'ReadyToImport_FormLabelKeys.csv' has been created!")
                logTheProgress(progressVal, progressMsg)

        
        ws.protection.enable()
        ws.protection.sort = False
        #ws.auto_filter.ref = "A1:D2"
        ws.protection.autoFilter = False
        ws.protection.formatColumns = False

        progressMsg = "Completed processing the '{}' sheet from the Workbook".format(wsName)
        logTheProgress(progressVal, progressMsg)
        
    progressVal += 1
    progressMsg = "Going to generate 'Ready To Import' Workbook and Translations configuration files."
    logTheProgress(progressVal, progressMsg)

    if len(modifiedSoups) > 0:
        print("Time to Save updated Excel file")
        translationsWb.save(importFilePath+"_WithChangeLog.xlsx")
        translationsWb.close()

    progressVal += 3
    progressMsg = "Updated Translations Workbook has been generated "
    logTheProgress(progressVal, progressMsg)

    progressLog += ("\n\n<<<<<<<<<<<***** Import Logs *****>>>>>>>>>>>\n\n"
            +importLogs)
    timestamp = time.strftime("%a_%d%b_%Y_%Hh%Mm%Ss", time.localtime())
    importLogFileName = "ImportLog_{}.log".format(timestamp)
    with open(fileSaveDir+"\\"+importLogFileName, "wt", encoding="utf-8") as file:
        file.write(importLogs)
    progressVal += 2
    progressMsg = "Import log file \"{}\" has been created.".format(importLogFileName)
    logTheProgress(progressVal, progressMsg)

    if len(modifiedSoups) > 0:
        progressIncr = 40/len(modifiedSoups)
    for soup in modifiedSoups:
        fileName = getSFSpecificName(soup)
        #updatedXML = soup.prettify(formatter=None)
        updatedXML = str(soup)
        readyToImportFileName = "ReadyToImport_{}.xml".format(fileName)
        with open(fileSaveDir+"\\"+readyToImportFileName, "w", encoding="utf-8") as file:
            file.write(updatedXML)
        progressVal += progressIncr
        progressMsg = ("'Ready to Import' configurations file {} " +
            "been generated. You can now import it to the SF instance!").format(readyToImportFileName)
        logTheProgress(progressVal, progressMsg)

    progressMsg=("Updated translations workbook and configuration files have "+
        "been generated and are ready to be imported!")
    logTheProgress(100, progressMsg)
    os.startfile(fileSaveDir)
    
    importExecuteBtn["state"] = "disabled"
    importExecuteBtn["image"] = doneIcon
    importExecuteBtn["compound"] = tk.LEFT
    # Proceed with update of Picklists
    #updatePicklistsLabels()

    # Time for update of Object Definitions

    # Update other objects
    

def readConfigXMLFiles():
    global soups
    global soupsDict
    global fileSaveDir
    global xmlFiles
    global translatableTags
    global fileOpenBtn

    xmlFiles = filedialog.askopenfilenames(initialdir=os.getcwd(), title="Open XML File", 
                                        filetypes=[("XML File","*.xml")])
    
    xmlFiles = list(xmlFiles)
    # New Lines of code added on 20.06.2023 to add standard translations from standard SAP SF DMs #
    yesNo = showMessage("Process Standard Data Models", "Do you want to add Standard translations from SAP for missing labels?", messagebox.YESNO)

    if yesNo == True:
        xmlFiles.append(os.path.join(pathToThisApp, stdSDMFileName))
        xmlFiles.append(os.path.join(pathToThisApp, stdCSF_SDMFileName))
        xmlFiles.append(os.path.join(pathToThisApp, stdCDMFileName))
        xmlFiles.append(os.path.join(pathToThisApp, stdCSF_CDMFileName))
    
    progressIncrement = 95
    filesCount = len(xmlFiles)
    print(filesCount)
    if(filesCount > 0):
        logTheProgress(1, "Extracting the content of the XML Files")
        progressIncrement = 95/filesCount
    
    progVal = 0
    tagNamesPerXMLFile = {}
    fileName = ""
    soup = None
    for xmlFile in xmlFiles:
        appUI.update_idletasks()
        xmlFile = xmlFile.replace("/","\\")
        fileName = xmlFile[xmlFile.rfind("\\")+1:]
        fileSaveDir = xmlFile[:xmlFile.rfind("\\")]
        
        xmlContent = ""
        with open(xmlFile, encoding="utf8") as f:
            xmlContent = f.read()
        
        if xmlContent.find("<![CDATA[") == -1:
            soup = BeautifulSoup(xmlContent, "xml")
        else:
            soup = BeautifulSoup(xmlContent, "html.parser")
        
        isStandardDM = False
        if fileSaveDir == pathToThisApp:
            isStandardDM = True
        xmlConfigName = getSFSpecificName(soup, isStandardDM)

        if soup.find("sf-pmreview") != None:
            xmlConfigName = fileName[:fileName.find(".xml")]
        
        if xmlConfigName == None:
            filesCount -= 1
            statusText = fileName + "could not be identified as a valid SF Data Model File or a valid Form Template file. Ignored for reading!"
        else:
            statusText = "Content of the Configuration File '"+xmlConfigName+"' has been extracted."
            soupsDict[xmlConfigName] = soup
            if not isStandardDM:
                soups.append(soup)
            tagNamesWithCount = []
            tagNames = []

            for tag in soup.findAll():
                tagName = tag.name
                if ((tagName != "role-name") and 
                    (tagName != "meta-grp-label") and
                    (tagName in ["instruction","label", "text", "default-rating", "unrated-rating"] 
                    or "-name" in tagName or "-label" in tagName or "-intro" in tagName 
                    or "-desc" in tagName)):
                    if tagName not in translatableTags:
                        translatableTags.append(tagName)

                if tagNames.count(tagName) == 0:
                    count = len(soup.findAll(tagName))
                    tagNames.append(tagName)
                    tagNamesWithCount.append(tagName + '(' + str(count)+')')
            
            tagNamesPerXMLFile[fileName] = tagNamesWithCount
        
        progVal += progressIncrement
        logTheProgress(progVal, statusText)
    
    progressMsg="Contents of {} SF data model files has been extracted! Connect to SF OData Service".format(filesCount)
    logTheProgress(100, progressMsg)
    
    #fileOpenBtn["text"] = "{} - ✅".format(fileOpenBtn["text"])
    if fileOpenBtn != None:
        fileOpenBtn["state"] = "disabled"
        fileOpenBtn["image"] = doneIcon
        fileOpenBtn["compound"] = tk.LEFT


def showExecutionStartButton():
    global exportExecuteBtn
    global importExecuteBtn
    progressMsg = "Ready"

    if isExportAction:
        exportExecuteBtn = ttk.Button(appUI, text="3. Now Execute Export", 
        command=executeExport)
        exportExecuteBtn.place(relx=0.25, rely=0.3, relwidth=0.2, relheight=0.05)
        progressMsg = "Translations can now be exported!"
    else:
        importExecuteBtn = ttk.Button(appUI, text="3. Now Execute Import", 
        command=executeImport)
        importExecuteBtn.place(relx=0.55, rely=0.3, relwidth=0.2, relheight=0.05)
        progressMsg = "Translations can now be imported!"

    logTheProgress(100, progressMsg)

def createSFODataService():
    global sfOdataService
    global connectApiBtn

    SERVICE_URL = ""
    companyId = ""
    username = ""
    password = ""
    connect = False

    if len(soups) == 0:
        showMessage("ERROR - Process Order Incorrect", 
        "Complete Step 1 before proceeding with this step.", 
        messagebox.ERROR)
        return

    if sfOdataService != None:
        if showMessage("Already an active OData connection - Use it?", "There is aleady an Active OData Session. "+
                    "Click 'Yes' to use the exiting connection or 'No' to connect afresh.",
                     messagebox.YESNO):
            connectApiBtn["state"] = "disabled"
            connectApiBtn["image"] = doneIcon
            connectApiBtn["compound"] = tk.LEFT
            showExecutionStartButton()
            return
    else:
        connect = showMessage("Connect to SF OData API or Skip", ("Extraction of translations fo some configuration items e.g. Picklists, Object Definitions,"+ 
                    "Org Objects requires connection to OData API. Do you want to extract or import translations "+
                    "for any of these configuration items? Click 'Yes' to enter OData API login credentials, 'No' to skip."),
                    messagebox.YESNO)
    
    if not connect:
        showExecutionStartButton()
        return

    session = requests.Session()
    msg = "Enter SF OData API Connection Details"
    title = "SF OData Connection"
    fieldNames = ["SF OData Service URL", "Company ID", "API Username", "Password"]
    ### refer here https://userapps.support.sap.com/sap/support/knowledge/en/2215682
    #defaultValues = ["https://apisalesdemo2.successfactors.eu/odata/v2", "SFPART034690", "AdminSK", ""]
    defaultValues = ["https://api12preview.sapsf.eu/odata/v2", "thyssenkagT9","10692666",""]
    #defaultValues = ["https://api012.successfactors.eu/odata/v2", "adidasD", "SFBACKGROUND", ""] # Also for hdiTest
    #defaultValues = ["https://api12preview.sapsf.eu/odata/v2", "deutschelT2","AdminSK",""]
    #defaultValues = ["https://api2.successfactors.eu/odata/v2", "claasD","ODataAPI_ApTrans","ApTrans@2020#3"]
    #defaultValues = ["https://api55preview.sapsf.eu/odata/v2", "dekrasedev","SFBACKGROUND","Dekra2021!"]
    fieldValues = gui.multpasswordbox(msg, title, fieldNames, defaultValues) 
    if fieldValues is None:
        sys.exit(0)
    # make sure that none of the fields were left blank
    while 1:
        errmsg = ""
        for i, name in enumerate(fieldNames):
            if fieldValues[i].strip() == "":
                errmsg += "{} is a required field.\n\n".format(name)
        if errmsg == "":
            SERVICE_URL = fieldValues[0]
            companyId = fieldValues[1]
            username = fieldValues[2]
            password = fieldValues[3]
            break # no problems found
        fieldValues = gui.multpasswordbox(errmsg, title, fieldNames, fieldValues)
        if fieldValues is None:
            break

    session.auth = (username+'@'+companyId, password)
    
    progressMsg="Establishing connection to SF OData Service..."
    logTheProgress(10, progressMsg)
    try:
        sfOdataService = pyodata.Client(SERVICE_URL, session)
        
        connectApiBtn["state"] = "disabled"
        connectApiBtn["image"] = doneIcon
        connectApiBtn["compound"] = tk.LEFT

        appUI.update()
        progressMsg="Connection to SF OData Service has been established!"
        logTheProgress(80, progressMsg)
        showExecutionStartButton()
        
    except pyodata.exceptions.HttpError:
        value = sys.exc_info()
        progress['value'] = 90
        progressMsg = "Connection to SF OData Service could not be established. Please retry!"
        showMessage("Connection Failed","The entered connection details are incorrect."+ 
            "Please make correction in accordance with the error "+
            "message recieved as below and then retry:\n\n"+str(value), 
            messagebox.ERROR)
        logTheProgress(100, progressMsg)
        #Retry
        createSFODataService()

def startExport():
    global isExportAction
    global fileOpenBtn
    global connectApiBtn

    isExportAction = True
    fileOpenBtn = ttk.Button(appUI, text="1. Open Configuration XML File(s)", command=readConfigXMLFiles)
    fileOpenBtn.place(relx=0.25, rely=0.1, relwidth=0.2, relheight=0.05)
    
    connectApiBtn = ttk.Button(appUI, text="2. Connect To SF OData Service", command=createSFODataService)
    connectApiBtn.place(relx=0.25, rely=0.2, relwidth=0.2, relheight=0.05)
  

def startImport():
    global translationsWb
    global importFilePath
    global fileOpenBtn
    global connectApiBtn
    global isExportAction

    isExportAction = False

    importFilePath = gui.fileopenbox("Browse to the updated Translations Workbook.", 
                    "Open Translations Workbook",fileSaveDir+"\\*.xlsx",["*.xlsx"])
    
    #read and check if the file is valid or not
    if importFilePath == None :
        return
    
    translationsWb = load_workbook(importFilePath)
    sheetNames = translationsWb.sheetnames
    matches = [x for x in sheetNames 
        if x in [sheetNameForPL, "FOAndConfigData", sheetNameForPM, sheetNameForGM] 
        or x.startswith("DataModel") or x.startswith("ObjectDefinitions")]
    if len(matches) == 0:
        showMessage("Invalid Translations Workbook!",
        "The excel file you uploaded is not a valid Translations file.")
        startImport()
    dmFiles = [x for x in sheetNames if x.startswith("DataModel")]
    
    text = "1. Browse to Latest EC Data Model File(s)"
    if sheetNameForPM in sheetNames:
        if len(dmFiles) == 0:
            text = "1. Browse to Latest PMGM Template File(s)"
        else:
            text = "1. Browse to Latest PMGM Template & Data Model File(s)"
    
    fileOpenBtn = ttk.Button(appUI, text=text, command=readConfigXMLFiles)
    fileOpenBtn.place(relx=0.55, rely=0.1, relwidth=0.2, relheight=0.05)

    connectApiBtn = ttk.Button(appUI, text="2. Connect SF OData Service", command=createSFODataService)
    connectApiBtn.place(relx=0.55, rely=0.2, relwidth=0.2, relheight=0.05)

    
def showAbout():
    showMessage("About {}".format(appName),"{} Version {}".format(appName, versionNum))

def showHelp():
    showMessage("{} - HELP".format(appName), ("{} is a acceleration application to easily manage translations of various translatable UI elements in SAP SuccessFactors."+
            "\nIt hugely reduces the manual effort needed in documenting the translations workbook as well as in preparing the configuration elements for importing translations")
                .format(appName))

def viewProgressLog():
    gui.textbox(msg="Following are the activities you have performed till now", 
    title = "{} - Progress Log".format(appName),
    text=progressLog)

def getAllUsedRules():
    if len(soups) == 0:
        readConfigXMLFiles()
    rulesWithReferences = findInDataModelReferencesOf("trigger-rule")
    ids = rulesWithReferences.get("ids")
    references = rulesWithReferences.get("references")
    resultToDisplay = "\n".join("{}: {}".format(ids[n], references[n] ) for n in range(len(ids)-1))
    gui.textbox("List of all Picklists Used in Data Model", "List of all Picklists Used in Data Model", resultToDisplay)
    # TODO #

def getAllUsedPicklists():
    gui.textbox("List of all Picklists Used in Data Model","List of all Picklists Used in Data Model & Object Definitions", "\n".join(picklistIds))
    # TODO #

menubar = tk.Menu(appUI)

actionMenu = tk.Menu(menubar, tearoff=False)
actionMenu.add_command(label="Export Translations", command=startExport)
actionMenu.add_command(label="Generate Import Configs For Translation", command=startImport)
actionMenu.add_command(label="Get List of all Picklists used", command=getAllUsedPicklists)
actionMenu.add_command(label="Get List of all Rules used", command=getAllUsedRules)
actionMenu.add_command(label="Reset Import/Export Progress", command=resetExecutionProgress)
menubar.add_cascade(label="Actions", menu=actionMenu)

aboutMenu = tk.Menu(menubar, tearoff=False)
aboutMenu.add_command(label="Using TREXIMA", command=showHelp)
aboutMenu.add_command(label="About TREXIMA", command=showAbout)
aboutMenu.add_command(label="Quit TREXIMA", command=appUI.quit)
menubar.add_cascade(label="About", menu=aboutMenu)

appUI.config(menu=menubar)

exportBtn = ttk.Button(appUI, text="Export Translations", command=startExport)
exportBtn.place(relx=0.25, rely=0.01, relwidth=0.2, relheight=0.05)

importBtn = ttk.Button(appUI, text="Generate Import Configs For Translation", command=startImport)
importBtn.place(relx=0.55, rely=0.01, relwidth=0.2, relheight=0.05)

viewProgressLogBtn = ttk.Button(appUI, text="View Progress Log", command=viewProgressLog)
viewProgressLogBtn.place(relx=0.40, rely=0.90, relwidth=0.2, relheight=0.05)

progress['maximum'] = 100
progress.place(relx=0.10, rely=0.80, relwidth=0.8, relheight=0.05)
progressMsg = ("{} is ready for export or import of applicable translations" +
                " configuration artifacts!").format(appName)
logTheProgress(0, progressMsg)

appUI.mainloop()


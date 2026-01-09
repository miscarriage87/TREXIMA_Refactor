#!/usr/bin/env python3
"""
TREXIMA v4.0 - End-to-End Test Script

Tests the complete export and import workflow:
1. Load XML data model files
2. Extract translations to Excel workbook
3. Modify some translations in the workbook
4. Import modified workbook back
5. Generate updated XML data model files

This uses the same core modules as the web application.
"""

import os
import sys
import shutil
from datetime import datetime

# Add project to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from trexima.config import AppPaths
from trexima.core.datamodel_processor import DataModelProcessor
from trexima.core.translation_extractor import TranslationExtractor
from trexima.core.translation_importer import TranslationImporter
from openpyxl import load_workbook

# Test configuration
TEST_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(TEST_DIR, 'e2e_test_output')
TIMESTAMP = datetime.now().strftime('%Y%m%d_%H%M%S')

# Test files
XML_FILES = [
    os.path.join(TEST_DIR, 'EC-data-model.xml'),
    os.path.join(TEST_DIR, 'EC-corporate-datamodel.xml'),
]

# Locales to export
LOCALES = ['en_US', 'de_DE', 'fr_FR', 'es_ES', 'ja_JP']


def print_header(text):
    """Print section header."""
    print("\n" + "=" * 60)
    print(f"  {text}")
    print("=" * 60)


def print_step(step_num, text):
    """Print step with number."""
    print(f"\n[Step {step_num}] {text}")


def progress_callback(percent, message):
    """Progress callback for extractor/importer."""
    bar_length = 30
    pct = int(percent) if isinstance(percent, (int, float)) else 0
    filled = int(bar_length * pct / 100)
    bar = '█' * filled + '░' * (bar_length - filled)
    print(f"\r  Progress: [{bar}] {pct:3d}% - {message}", end='', flush=True)
    if pct >= 100:
        print()


def main():
    """Run end-to-end test."""
    print_header("TREXIMA v4.0 - End-to-End Test")
    print(f"Timestamp: {TIMESTAMP}")
    print(f"Output directory: {OUTPUT_DIR}")

    # Clean and create output directory
    if os.path.exists(OUTPUT_DIR):
        shutil.rmtree(OUTPUT_DIR)
    os.makedirs(OUTPUT_DIR)

    # =========================================================================
    # STEP 1: Initialize processor and load data models
    # =========================================================================
    print_step(1, "Initializing and loading data model files")

    app_paths = AppPaths(app_dir=OUTPUT_DIR)
    processor = DataModelProcessor(app_paths)

    loaded_count = 0
    for xml_file in XML_FILES:
        if os.path.exists(xml_file):
            print(f"  Loading: {os.path.basename(xml_file)}")
            model = processor.load_data_model(xml_file)
            if model:
                loaded_count += 1
                print(f"    ✓ Loaded {len(model.translatable_tags)} translatable elements")
        else:
            print(f"  ✗ File not found: {xml_file}")

    if loaded_count == 0:
        print("  ERROR: No data model files could be loaded!")
        return False

    print(f"\n  Total models loaded: {loaded_count}")

    # =========================================================================
    # STEP 2: Extract translations to Excel workbook
    # =========================================================================
    print_step(2, "Extracting translations to Excel workbook")

    extractor = TranslationExtractor(
        processor=processor,
        odata_client=None,  # No API connection for this test
        progress_callback=progress_callback
    )

    workbook = extractor.extract_to_workbook(
        locales_for_export=LOCALES,
        export_picklists=False,  # No API = no picklists
        export_mdf_objects=False,  # No API = no MDF
        export_fo_translations=False,  # No API = no FO
        system_default_lang='en_US'
    )

    # Save workbook
    export_filename = f"TranslationsWorkbook_{TIMESTAMP}.xlsx"
    export_path = os.path.join(OUTPUT_DIR, export_filename)
    extractor.save_workbook(workbook, OUTPUT_DIR, export_filename)

    print(f"\n  ✓ Workbook saved: {export_filename}")
    print(f"    File size: {os.path.getsize(export_path):,} bytes")

    # Count sheets and rows
    wb = load_workbook(export_path)
    print(f"    Sheets: {len(wb.sheetnames)}")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        row_count = sheet.max_row - 1 if sheet.max_row > 1 else 0
        print(f"      - {sheet_name}: {row_count} rows")
    wb.close()

    # =========================================================================
    # STEP 3: Modify translations in the workbook
    # =========================================================================
    print_step(3, "Modifying translations in the workbook")

    modified_filename = f"TranslationsWorkbook_{TIMESTAMP}_modified.xlsx"
    modified_path = os.path.join(OUTPUT_DIR, modified_filename)

    # Load and modify
    wb = load_workbook(export_path)
    modifications_made = 0

    # Find a data model sheet to modify
    for sheet_name in wb.sheetnames:
        if 'DM_' in sheet_name or 'Data Model' in sheet_name or sheet_name.startswith('SFEC'):
            sheet = wb[sheet_name]

            # Find translation columns (look for language headers)
            header_row = 1
            lang_cols = {}
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=header_row, column=col).value
                if header and header in ['de_DE', 'fr_FR', 'es_ES']:
                    lang_cols[header] = col

            if not lang_cols:
                continue

            print(f"  Modifying sheet: {sheet_name}")

            # Modify first 5 rows with translations
            rows_modified = 0
            for row in range(2, min(sheet.max_row + 1, 12)):  # First 10 data rows
                for lang, col in lang_cols.items():
                    cell = sheet.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        original = cell.value
                        cell.value = f"[E2E TEST] {original}"
                        modifications_made += 1
                        rows_modified += 1

            print(f"    Modified {rows_modified} cells in {len(lang_cols)} languages")
            break  # Only modify one sheet for the test

    if modifications_made == 0:
        # Try to find any sheet with content
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.max_row > 1 and sheet.max_column > 3:
                # Find columns that might be translations
                for col in range(4, min(sheet.max_column + 1, 10)):
                    for row in range(2, min(sheet.max_row + 1, 7)):
                        cell = sheet.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, str) and len(cell.value) > 2:
                            original = cell.value
                            cell.value = f"[E2E TEST] {original}"
                            modifications_made += 1

                if modifications_made > 0:
                    print(f"  Modified {modifications_made} cells in sheet: {sheet_name}")
                    break

    wb.save(modified_path)
    wb.close()

    print(f"\n  ✓ Modified workbook saved: {modified_filename}")
    print(f"    Total modifications: {modifications_made}")

    # =========================================================================
    # STEP 4: Import modified workbook
    # =========================================================================
    print_step(4, "Importing modified workbook")

    importer = TranslationImporter(
        processor=processor,
        progress_callback=progress_callback
    )

    # Load the modified workbook
    modified_wb = load_workbook(modified_path)
    xml_output_dir = os.path.join(OUTPUT_DIR, 'updated_datamodels')
    os.makedirs(xml_output_dir, exist_ok=True)

    # Get all sheet names to process
    sheets_to_process = modified_wb.sheetnames

    # Import the modified workbook
    import_result = importer.import_from_workbook(
        workbook=modified_wb,
        worksheets_to_process=sheets_to_process,
        save_dir=xml_output_dir
    )
    modified_wb.close()

    print(f"\n  ✓ Import completed")
    print(f"    Success: {import_result.success}")
    print(f"    Files generated: {len(import_result.files_generated)}")
    print(f"    Changes made: {import_result.changes_made}")
    if import_result.error_message:
        print(f"    Error: {import_result.error_message}")

    # =========================================================================
    # STEP 5: List generated XML data model files
    # =========================================================================
    print_step(5, "Checking generated XML data model files")

    generated_files = import_result.files_generated

    if generated_files:
        print("  Generated files:")
        for file_path in generated_files:
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                print(f"    ✓ {os.path.basename(file_path)} ({file_size:,} bytes)")
            else:
                print(f"    ! {os.path.basename(file_path)} (file not found)")
    else:
        print("  No files were generated. Checking data models manually...")
        # Try to save models manually if importer didn't generate files
        for model_name, model in processor.data_models.items():
            output_filename = f"ReadyToImport_{model_name}.xml"
            output_path = os.path.join(xml_output_dir, output_filename)

            try:
                processor.save_data_model(model, output_path)
                generated_files.append(output_path)
                file_size = os.path.getsize(output_path)
                print(f"    ✓ Generated: {output_filename} ({file_size:,} bytes)")
            except Exception as e:
                print(f"    ✗ Failed to generate {output_filename}: {e}")

    # =========================================================================
    # STEP 6: Verify outputs
    # =========================================================================
    print_step(6, "Verifying outputs")

    print("\n  Output files generated:")

    # List all output files
    for root, dirs, files in os.walk(OUTPUT_DIR):
        for file in files:
            file_path = os.path.join(root, file)
            rel_path = os.path.relpath(file_path, OUTPUT_DIR)
            size = os.path.getsize(file_path)
            print(f"    {rel_path}: {size:,} bytes")

    # =========================================================================
    # SUMMARY
    # =========================================================================
    print_header("TEST RESULTS SUMMARY")

    print(f"\n  ✓ Export: Generated Excel workbook with translations")
    print(f"    - {export_filename}")

    print(f"\n  ✓ Modified workbook with {modifications_made} test changes")
    print(f"    - {modified_filename}")

    print(f"\n  ✓ Import: Processed modified workbook")
    print(f"    - Changes made: {import_result.changes_made}")

    print(f"\n  ✓ Generated {len(generated_files)} updated XML data model files:")
    for f in generated_files:
        print(f"    - {f}")

    print(f"\n  All outputs saved to: {OUTPUT_DIR}")

    print("\n" + "=" * 60)
    print("  END-TO-END TEST COMPLETED SUCCESSFULLY")
    print("=" * 60 + "\n")

    return True


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)

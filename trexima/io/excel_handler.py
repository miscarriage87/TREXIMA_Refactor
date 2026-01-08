"""
Excel Handler Module

Handles reading and writing of Excel workbooks.
"""

import os
import time
from typing import List, Optional, Dict, Any

import babel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    NamedStyle, Font, PatternFill, Border, Protection,
    GradientFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..config import (
    SHEET_NAME_PM,
    SHEET_NAME_GM,
    SHEET_NAME_PL,
    WORKBOOK_PASSWORD
)


class ExcelHandler:
    """Handles Excel workbook operations."""

    def __init__(self):
        # Border styles
        self.thin = Side(border_style="thin", color="000000")
        self.thick = Side(border_style="thick", color="000000")
        self.border = Border(
            top=self.thin, left=self.thin,
            right=self.thin, bottom=self.thin
        )
        self.header_cell_border = Border(
            top=self.thick, left=self.thick,
            right=self.thick, bottom=self.thick
        )

        # Font styles
        self.bold_font = Font(color="00000000", bold=True, size=12)

        # Fill styles
        self.grad_fill = GradientFill(
            degree=float(90), stop=("FFFFFF", "FED8B1")
        )
        self.grad_header_fill = GradientFill(
            degree=float(90), stop=("FFFFFF", "E8E8E8")
        )
        self.grad_edit_fill = GradientFill(
            degree=float(90), stop=("FFFFFF", "C7F6C7")
        )

    def create_workbook(self) -> Workbook:
        """Create a new workbook with standard styles."""
        workbook = Workbook()
        workbook.iso_dates = True
        self._setup_styles(workbook)
        return workbook

    def load_workbook(self, file_path: str) -> Workbook:
        """Load an existing workbook."""
        return load_workbook(file_path)

    def _setup_styles(self, workbook: Workbook):
        """Set up named styles for the workbook."""
        if "BoldCellStyle" not in workbook.named_styles:
            bold_style = NamedStyle(name="BoldCellStyle")
            bold_style.font = self.bold_font
            workbook.add_named_style(bold_style)

        if "HeadersStyle" not in workbook.named_styles:
            header_style = NamedStyle(name="HeadersStyle")
            header_style.font = self.bold_font
            header_style.fill = self.grad_header_fill
            header_style.border = self.border
            workbook.add_named_style(header_style)

        if "LockedCellStyle" not in workbook.named_styles:
            locked_style = NamedStyle(name="LockedCellStyle")
            locked_style.font = Font(color="00003F")
            locked_style.fill = self.grad_fill
            locked_style.border = self.border
            workbook.add_named_style(locked_style)

        if "EditableCellStyle" not in workbook.named_styles:
            editable_style = NamedStyle(name="EditableCellStyle")
            editable_style.fill = self.grad_edit_fill
            editable_style.border = self.border
            editable_style.protection = Protection(locked=False, hidden=False)
            workbook.add_named_style(editable_style)

    def create_sheets_per_lang(
        self,
        workbook: Workbook,
        sheet_name: str,
        langs: List[str],
        base_headers: List[str]
    ) -> List[str]:
        """
        Create worksheets for each language.

        Args:
            workbook: Workbook to add sheets to
            sheet_name: Base sheet name
            langs: List of language codes
            base_headers: Headers for the sheet

        Returns:
            List of created worksheet names
        """
        worksheets = []
        header_cols = len(base_headers)
        header_cell_id = f"{get_column_letter(header_cols + 1)}1"

        for lang_id in langs:
            worksheet_name = f"{sheet_name} ({lang_id})"
            separator = "-" if "-" in lang_id else "_"

            if worksheet_name not in worksheets:
                worksheet = workbook.create_sheet(worksheet_name)
                worksheets.append(worksheet_name)
                worksheet.append(base_headers)

                # Parse locale for display name
                parse_lang = lang_id
                if lang_id.startswith(f"bs{separator}"):
                    parse_lang = "bs"

                if lang_id != f"en{separator}DEBUG":
                    try:
                        locale_name = babel.Locale.parse(
                            parse_lang, sep=separator
                        ).english_name
                        worksheet[header_cell_id] = f"Label/Name in {locale_name}"
                    except:
                        worksheet[header_cell_id] = f"Label/Name in {lang_id}"
                else:
                    worksheet[header_cell_id] = "Label/Name in SF Debug Language"

        return worksheets

    def change_cell_style(
        self,
        cell,
        new_style_name: str
    ):
        """Change cell style while preserving bold formatting."""
        is_bold = cell.font.b if cell.font else False

        if cell.row == 1 and new_style_name == "EditableCellStyle" and cell.value:
            cell.style = "LockedCellStyle"
        else:
            cell.style = new_style_name

        if is_bold or cell.row == 1:
            cell.font = self.bold_font
        if cell.row == 1:
            cell.border = self.header_cell_border

    def append_as_header_row(
        self,
        worksheet: Worksheet,
        cell_values: List
    ):
        """Append a row with header styling."""
        if worksheet:
            worksheet.append(cell_values)
            last_row = worksheet[worksheet.max_row]
            for cell in last_row:
                cell.style = "BoldCellStyle"

    def prepare_and_save_workbook(
        self,
        workbook: Workbook,
        file_path: str
    ):
        """
        Prepare workbook with protection and formatting, then save.

        Args:
            workbook: Workbook to prepare
            file_path: Path to save the workbook
        """
        self._setup_styles(workbook)

        # Set workbook properties
        workbook.security.workbookPassword = WORKBOOK_PASSWORD
        workbook.security.lockStructure = True
        workbook.properties.title = "SF Translations Workbook - By TREXIMA"
        workbook.properties.category = "TranslationsWorkbook"
        workbook.properties.subject = "SF Translations Workbook"
        workbook.properties.creator = "TREXIMA by Sandeep Kumar (Deloitte)"

        # Format each worksheet
        for ws in workbook:
            self._format_worksheet(ws)

        workbook.save(file_path)

    def _format_worksheet(self, ws: Worksheet):
        """Apply formatting to a worksheet."""
        from datetime import datetime, date, time

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1

        # First pass: Clean ALL datetime objects in the worksheet (Excel doesn't support timezones)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (datetime, time)):
                    if cell.value.tzinfo is not None:
                        cell.value = cell.value.replace(tzinfo=None)

        # Second pass: Format first 10 rows and adjust column widths
        for row in ws:
            if row[0].row == 10:
                break
            for cell in row:
                if cell.row == 1:
                    cell.style = "HeadersStyle"
                col_width = ws.column_dimensions[cell.column_letter].width
                cells_needed_width = len(str(cell.value)) if cell.value else 0
                if col_width <= cells_needed_width + 2:
                    ws.column_dimensions[cell.column_letter].width = min(
                        35, cells_needed_width + 2
                    )

        # Set protection
        ws.protection.password = WORKBOOK_PASSWORD
        ws.protection.sheet = True
        ws.protection.sort = False
        ws.auto_filter.ref = "A1:D2"
        ws.protection.autoFilter = False
        ws.protection.formatColumns = False

        # Determine protected columns based on sheet type
        protected_cols = ["A", "B"]
        freeze_cells_range = ws["C2"]

        if ws.title == SHEET_NAME_PL or ws.title.startswith("DataModel"):
            protected_cols.extend(["C", "D"])
            freeze_cells_range = ws["E2"]
        elif not ws.title.startswith("ObjectDefinitions"):
            protected_cols.extend(["C", "D"])
            freeze_cells_range = ws["E2"]

        ws.freeze_panes = freeze_cells_range
        locked_cols_count = len(protected_cols)

        # Apply locked style to protected columns
        for col in ws.iter_cols(
            min_col=1, max_col=locked_cols_count, min_row=1
        ):
            for cell in col:
                self.change_cell_style(cell, "LockedCellStyle")

        # Apply editable style to other columns
        for col in ws.iter_cols(
            min_col=locked_cols_count + 1,
            max_col=ws.max_column + 2,
            min_row=1
        ):
            for cell in col:
                self.change_cell_style(cell, "EditableCellStyle")

        # Special handling for PM sheet
        if ws.title == SHEET_NAME_PM:
            for cell in ws["F"]:
                self.change_cell_style(cell, "LockedCellStyle")

    def generate_export_filename(self, save_dir: str) -> str:
        """Generate a timestamped export filename."""
        timestamp = time.strftime("%a_%d%b_%Y_%Hh%Mm%Ss", time.localtime())
        return os.path.join(save_dir, f"SF_EC_Translations_{timestamp}.xlsx")

    def get_language_header_name(self, lang_id: str, separator: str = "_") -> str:
        """Get display name for a language header."""
        parse_lang = lang_id
        if lang_id.startswith(f"bs{separator}"):
            parse_lang = "bs"

        if lang_id == f"en{separator}DEBUG":
            return "Label SF Debug (en-DEBUG)"

        try:
            locale_name = babel.Locale.parse(parse_lang, sep=separator).english_name
            return f"Label in {locale_name} ({lang_id})"
        except:
            return f"Label in {lang_id}"

    def validate_translations_workbook(
        self,
        workbook: Workbook
    ) -> bool:
        """
        Validate that a workbook is a valid translations workbook.

        Args:
            workbook: Workbook to validate

        Returns:
            True if valid
        """
        sheet_names = workbook.sheetnames
        valid_prefixes = [
            SHEET_NAME_PL,
            "FOAndConfigData",
            SHEET_NAME_PM,
            SHEET_NAME_GM,
            "DataModel",
            "ObjectDefinitions"
        ]

        for name in sheet_names:
            for prefix in valid_prefixes:
                if name == prefix or name.startswith(prefix):
                    return True

        return False

    def get_datamodel_sheets(self, workbook: Workbook) -> List[str]:
        """Get list of DataModel sheet names."""
        return [
            name for name in workbook.sheetnames
            if name.startswith("DataModel")
        ]

    def add_change_log_column(
        self,
        worksheet: Worksheet,
        header_text: str = "Change Log Identified During Import"
    ) -> int:
        """
        Add a change log column to a worksheet.

        Args:
            worksheet: Worksheet to modify
            header_text: Header text for the column

        Returns:
            Column number of the change log column
        """
        # Find the last non-empty column
        col_num = 1
        max_col_num = worksheet.max_column
        for cell in worksheet[1]:
            if cell.value is None or cell.value == "":
                max_col_num = col_num
                break
            col_num += 1

        # Add header
        change_log_header = worksheet.cell(row=1, column=max_col_num)
        change_log_header.value = header_text
        change_log_header.style = "HeadersStyle"
        worksheet.column_dimensions[change_log_header.column_letter].width = 75

        return max_col_num

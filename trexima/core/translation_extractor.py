"""
Translation Extractor Module

Extracts translations from SF data models and exports to Excel.
"""

import os
import time
from typing import List, Optional, Dict, Any, Callable

import babel
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..config import (
    SHEET_NAME_PM,
    SHEET_NAME_GM,
    SHEET_NAME_PL,
    EMPLOYEE_PROFILE_TAGS,
    TAGS_TO_BE_IGNORED,
    HIGHLIGHT_TAGS,
    KEYWORD_SAP_STANDARD
)
from ..models.datamodel import DataModel, DataModelType, ExportResult
from ..io.xml_handler import XMLHandler
from ..io.excel_handler import ExcelHandler
from .datamodel_processor import DataModelProcessor
from .odata_client import ODataClient


class TranslationExtractor:
    """Extracts translations from SF configurations."""

    def __init__(
        self,
        processor: DataModelProcessor,
        odata_client: Optional[ODataClient] = None,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ):
        self.processor = processor
        self.odata_client = odata_client
        self.xml_handler = XMLHandler()
        self.excel_handler = ExcelHandler()
        self.progress_callback = progress_callback

        # State
        self.picklist_ids: List[str] = []
        self.picklist_references: List[str] = []
        self.label_keys_dict: Dict[str, Dict] = {}
        self.label_keys_headers: List[str] = []
        self.active_countries: List[str] = []

    def _log_progress(self, percent: int, message: str):
        """Log progress if callback is set."""
        if self.progress_callback:
            self.progress_callback(percent, message)

    def extract_to_workbook(
        self,
        locales_for_export: List[str],
        # Picklists - separated by type
        export_mdf_picklists: bool = False,
        export_legacy_picklists: bool = False,
        # MDF Objects
        export_mdf_objects: bool = True,
        mdf_objects_filter: Optional[List[str]] = None,
        # FO Translations
        export_fo_translations: bool = True,
        fo_objects_filter: Optional[List[str]] = None,
        fo_translation_types_filter: Optional[List[str]] = None,
        # Other options
        picklist_from_csv: Optional[str] = None,
        remove_html_tags: bool = False,
        system_default_lang: str = "en_US"
    ) -> Workbook:
        """
        Extract translations to an Excel workbook.

        Args:
            locales_for_export: List of locale codes to export
            export_mdf_picklists: Whether to export MDF (PickListV2) picklists
            export_legacy_picklists: Whether to export legacy picklists
            export_mdf_objects: Whether to export MDF object definitions
            mdf_objects_filter: List of MDF object IDs to include (empty = all)
            export_fo_translations: Whether to export foundation objects
            fo_objects_filter: List of FO object IDs to include (empty = all)
            fo_translation_types_filter: List of FO translation type IDs to include
            picklist_from_csv: Path to picklist CSV file
            remove_html_tags: Whether to remove HTML tags from labels
            system_default_lang: System default language

        Returns:
            Workbook with exported translations
        """
        workbook = self.excel_handler.create_workbook()
        progress = 0

        # Determine if any picklists should be exported
        export_any_picklists = export_mdf_picklists or export_legacy_picklists

        # Collect picklist references if needed
        if export_any_picklists:
            self._log_progress(5, "Collecting picklist references from data models...")
            self.picklist_ids, self.picklist_references = (
                self.processor.find_picklist_references()
            )
            progress = 10

        # Export MDF object definitions
        if export_mdf_objects and self.odata_client and self.odata_client.is_connected:
            self._log_progress(progress, "Extracting MDF object definitions...")
            self._export_mdf_objects(
                workbook, locales_for_export, system_default_lang,
                mdf_objects_filter=mdf_objects_filter,
                fo_objects_filter=fo_objects_filter
            )
            progress = 20

        # Export picklists
        if export_any_picklists:
            self._log_progress(progress, "Extracting picklists...")
            if picklist_from_csv:
                self._export_picklists_from_csv(workbook, picklist_from_csv)
            elif self.odata_client and self.odata_client.is_connected:
                self._export_picklists_from_api(
                    workbook, locales_for_export, system_default_lang,
                    include_mdf=export_mdf_picklists,
                    include_legacy=export_legacy_picklists
                )
            progress = 40

        # Export foundation objects
        if export_fo_translations and self.odata_client and self.odata_client.is_connected:
            self._log_progress(progress, "Extracting foundation objects...")
            self._export_foundation_objects(
                workbook, locales_for_export,
                fo_translation_types_filter=fo_translation_types_filter
            )
            progress = 50

        # Export data model translations
        self._log_progress(progress, "Extracting data model translations...")
        self._export_datamodel_translations(
            workbook,
            locales_for_export,
            remove_html_tags,
            system_default_lang
        )

        # Remove default sheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        self._log_progress(100, "Export complete!")
        return workbook

    def _export_mdf_objects(
        self,
        workbook: Workbook,
        locales: List[str],
        default_lang: str,
        mdf_objects_filter: Optional[List[str]] = None,
        fo_objects_filter: Optional[List[str]] = None
    ):
        """
        Export MDF object definitions.

        Args:
            workbook: Target workbook
            locales: List of locale codes
            default_lang: Default language code
            mdf_objects_filter: List of custom MDF object IDs to include (empty = all)
            fo_objects_filter: List of FO object IDs to include (empty = all)
        """
        if not self.odata_client or not self.odata_client.is_connected:
            return

        base_headers = ["Object", "Id"]
        self.excel_handler.create_sheets_per_lang(
            workbook, "ObjectDefinitions", locales, base_headers
        )

        # Standard MDF entities (always included unless filtered)
        standard_entities = [
            "Position", "PaymentInformationDetailV3", "FOCompany",
            "FOBusinessUnit", "FODivision", "FODepartment", "FOJobCode",
            "FOJobFunction", "FOCostCenter", "FOPayGroup"
        ]

        # Filter FO objects if filter is provided
        if fo_objects_filter:
            fo_filter_set = set(fo_objects_filter)
            mdf_entities = [
                e for e in standard_entities
                if not e.startswith("FO") or e in fo_filter_set
            ]
        else:
            mdf_entities = standard_entities.copy()

        # Add custom objects
        all_entities = self.odata_client.get_all_entity_names()
        custom_objs = [e for e in all_entities if e.startswith("cust_")]

        # Filter custom MDF objects if filter is provided
        if mdf_objects_filter:
            mdf_filter_set = set(mdf_objects_filter)
            custom_objs = [e for e in custom_objs if e in mdf_filter_set]

        mdf_entities.extend(custom_objs)

        for object_name in mdf_entities:
            obj_id = object_name
            if object_name == "FOCompany":
                obj_id = "LegalEntity"

            for lang in locales:
                ws_name = f"ObjectDefinitions ({lang})"
                if ws_name not in workbook.sheetnames:
                    continue

                ws = workbook[ws_name]
                metadata = self.odata_client.get_mdf_object_metadata(object_name, lang)

                if metadata:
                    entity_sets = metadata.find_all("EntitySet")
                    if entity_sets:
                        entity_set = entity_sets[0]
                        ws.append([
                            entity_set.get("Name"),
                            obj_id,
                            entity_set.get("sap:label")
                        ])
                        last_row = ws[ws.max_row]
                        for cell in last_row:
                            cell.style = "BoldCellStyle"

                        # Add properties
                        properties = metadata.find_all("Property")
                        for prop in properties:
                            if prop.get("sap:visible") == "true":
                                ws.append([
                                    obj_id,
                                    prop.get("Name"),
                                    prop.get("sap:label")
                                ])

                                # Track picklists
                                picklist_id = prop.get("sap:picklist")
                                if picklist_id and picklist_id not in self.picklist_ids:
                                    self.picklist_ids.append(picklist_id)
                                    self.picklist_references.append(
                                        f"MDF Object ({obj_id}) -> Field ({prop.get('Name')})"
                                    )

    def _export_picklists_from_api(
        self,
        workbook: Workbook,
        locales: List[str],
        default_lang: str,
        include_mdf: bool = True,
        include_legacy: bool = True
    ):
        """
        Export picklists from OData API.

        Args:
            workbook: Target workbook
            locales: List of locale codes
            default_lang: Default language code
            include_mdf: Whether to include MDF (PickListV2) picklists
            include_legacy: Whether to include legacy picklists
        """
        if not self.odata_client or not self.odata_client.is_connected:
            return

        picklist_items = []
        batch_size = 10

        # Fetch legacy picklists if requested
        if include_legacy:
            # Check for migrated legacy picklists
            migrated_count = self.odata_client.get_migrated_legacy_picklist_count()

            if migrated_count == 0:
                # Fetch legacy picklists only if not migrated
                total_legacy = self.odata_client.get_picklist_count("legacy")
                offset = 0
                while offset < total_legacy:
                    items = self.odata_client.get_legacy_picklists(batch_size, offset)
                    picklist_items.extend(items)
                    offset += batch_size

        # Fetch MDF picklists if requested
        if include_mdf:
            total_mdf = self.odata_client.get_picklist_count("mdf")
            offset = 0
            while offset < total_mdf:
                items = self.odata_client.get_mdf_picklists(batch_size, offset)
                picklist_items.extend(items)
                offset += batch_size

        # Create worksheet
        headers = ["Reference", "Picklist Id", "Option's Unique Code", "Option ID"]
        for locale in locales:
            headers.append(f"Option Label ({locale})")

        ws = workbook.create_sheet(SHEET_NAME_PL)
        ws.append(headers)

        # Process picklists
        for picklist_item in picklist_items:
            is_mdf = False
            try:
                picklist_item.__getattr__("values")
                picklist_id = picklist_item.id
                options_prop = "values"
                is_mdf = True
            except AttributeError:
                picklist_id = picklist_item.picklistId
                options_prop = "picklistOptions"

            # Get references
            references = ""
            while picklist_id in self.picklist_ids:
                idx = self.picklist_ids.index(picklist_id)
                self.picklist_ids.pop(idx)
                ref = self.picklist_references.pop(idx)
                references += ref

            if not references:
                continue

            # Add options
            for option in picklist_item.__getattr__(options_prop):
                external_code = option.externalCode
                row_data = [references, picklist_id, external_code]
                labels = []

                if is_mdf:
                    option_id = option.optionId
                    for locale in locales:
                        label = option.__getattr__(f"label_{locale}")
                        if locale == default_lang:
                            labels.insert(0, label)
                        else:
                            labels.append(label)
                else:
                    option_id = option.id
                    for locale in locales:
                        for pl_label in option.picklistLabels:
                            if locale == pl_label.locale:
                                label = pl_label.label
                                if locale == default_lang:
                                    labels.insert(0, label)
                                else:
                                    labels.append(label)
                                break

                row_data.append(option_id)
                row_data.extend(labels)
                ws.append(row_data)

    def _export_picklists_from_csv(
        self,
        workbook: Workbook,
        csv_path: str
    ):
        """Export picklists from CSV file."""
        from ..io.csv_handler import CSVHandler

        csv_handler = CSVHandler()
        values_matrix = csv_handler.read_csv_as_matrix(csv_path)

        ws = workbook.create_sheet(SHEET_NAME_PL)
        ws.column_dimensions[get_column_letter(1)].width = 75

        headers = values_matrix[0]
        col_num = 0
        for header in headers:
            col_num += 1
            if col_num > 1 and header not in ["id", "values.externalCode"]:
                if not header.startswith("values.label."):
                    ws.column_dimensions[get_column_letter(col_num)].hidden = True

        row_num = 0
        picklist_id = ""
        references = ""

        for row in values_matrix:
            row_num += 1
            if picklist_id != row[1]:
                references = ""
                picklist_id = row[1]

                while picklist_id in self.picklist_ids:
                    idx = self.picklist_ids.index(picklist_id)
                    self.picklist_ids.pop(idx)
                    ref = self.picklist_references.pop(idx)
                    if ref not in references:
                        references += ref

            if row_num < 3:
                row[0] = "References in EC"
            else:
                row[0] = references

            ws.append(row)

            if row_num > 2 and not references:
                ws.row_dimensions[row_num].hidden = True

    def _export_foundation_objects(
        self,
        workbook: Workbook,
        locales: List[str],
        fo_translation_types_filter: Optional[List[str]] = None
    ):
        """
        Export foundation object translations.

        Args:
            workbook: Target workbook
            locales: List of locale codes
            fo_translation_types_filter: List of FO translation type IDs to include
                                         (empty = all). Maps to objects via FO_TRANSLATION_TYPES.
        """
        if not self.odata_client or not self.odata_client.is_connected:
            return

        ws = workbook.create_sheet("FOAndConfigData")
        headers = ["FO/Config Name", "Code", "Field", "Default Label"]

        for lang in locales:
            try:
                locale_name = babel.Locale.parse(lang).english_name
                headers.append(f"Label in {locale_name}")
            except (ValueError, AttributeError, babel.UnknownLocaleError):
                headers.append(f"Label in {lang}")

        ws.append(headers)

        entity_names = self.odata_client.get_all_entity_names()
        fo_entities = [
            e for e in entity_names
            if (e.startswith("FO") or e.startswith("cust_"))
            and not e.startswith("FOW")
            and not e.endswith("DEFLT")
        ]

        # Filter FO entities based on fo_translation_types_filter
        if fo_translation_types_filter:
            from trexima.web.constants import FO_TRANSLATION_TYPES
            allowed_objects = set()
            for fo_type in FO_TRANSLATION_TYPES:
                if fo_type['id'] in fo_translation_types_filter:
                    allowed_objects.add(fo_type['object'])
            fo_entities = [e for e in fo_entities if e in allowed_objects]

        # Add common objects
        objects_to_export = ["AlertMessage", "CustomPayType"]
        objects_to_export.extend(fo_entities)

        for obj_name in objects_to_export:
            trans_props, trans_fields = self.odata_client.get_translatable_properties(
                obj_name, locales
            )
            is_legacy_fo = self.odata_client.has_name_translation_nav(obj_name)

            try:
                if trans_props:
                    objects = self.odata_client.get_foundation_objects(obj_name)
                elif is_legacy_fo:
                    objects = self.odata_client.get_foundation_objects(
                        obj_name,
                        ["nameTranslationNav", "descriptionTranslationNav"]
                    )
                else:
                    continue

                metadata = self.odata_client.get_entity_metadata(obj_name)
                key_prop = metadata.key_proprties[0].name if metadata else "externalCode"

                for obj in objects:
                    try:
                        key = obj.__getattr__(key_prop)
                    except (AttributeError, KeyError):
                        continue

                    if is_legacy_fo:
                        # Handle legacy FO with translation nav
                        self._add_legacy_fo_row(ws, obj_name, key, obj, locales, "name")
                        self._add_legacy_fo_row(ws, obj_name, key, obj, locales, "description")
                    else:
                        # Handle MDF objects
                        for field in trans_fields:
                            row = [obj_name, key, field]
                            try:
                                default_val = obj.__getattr__(f"{field}_defaultValue")
                            except (AttributeError, KeyError):
                                default_val = None

                            if default_val:
                                row.append(default_val)
                            else:
                                row.append(obj.__getattr__(field))

                            for lang in locales:
                                row.append(obj.__getattr__(f"{field}_{lang}"))

                            ws.append(row)

            except Exception as e:
                print(f"Error processing {obj_name}: {e}")

    def _add_legacy_fo_row(
        self,
        ws,
        obj_name: str,
        key: str,
        obj,
        locales: List[str],
        field: str
    ):
        """Add a row for legacy FO translation."""
        nav_name = f"{field}TranslationNav"
        row = [obj_name, key, field]

        try:
            trans_nav = obj.__getattr__(nav_name)
            if trans_nav:
                row.append(trans_nav.__getattr__("value_defaultValue"))
                for lang in locales:
                    try:
                        row.append(trans_nav.__getattr__(f"value_{lang}"))
                    except (AttributeError, KeyError):
                        row.append("")
            else:
                row.append(obj.__getattr__(field))
                for _ in locales:
                    row.append("")
        except (AttributeError, KeyError, TypeError):
            row.append("")
            for _ in locales:
                row.append("")

        ws.append(row)

    def _export_datamodel_translations(
        self,
        workbook: Workbook,
        locales: List[str],
        remove_html_tags: bool,
        default_lang: str
    ):
        """Export data model translations."""
        xml_langs = [x.replace("_", "-") for x in locales]

        # Create sheets for PM/GM if needed
        worksheets = []
        headers = ["Section", "Element/Subsection", "Field Id", "Default Label"]

        if self.processor.is_pmgm_included:
            headers = [
                "Translation Type", "Template Name", "Section/Element/Subsection",
                "Translatable Item/Field", "Default Label", "Label Key"
            ]

            for lang_id in xml_langs:
                lang = lang_id.replace("-", "_")
                parse_lang = lang_id if not lang_id.startswith("bs-") else "bs"
                if lang_id != "en-DEBUG":
                    try:
                        locale_name = babel.Locale.parse(parse_lang, sep="-").english_name
                        headers.append(f"Label in {locale_name} ({lang})")
                    except (ValueError, AttributeError, babel.UnknownLocaleError):
                        headers.append(f"Label in {lang}")
                else:
                    headers.append("Label SF Debug (en-DEBUG)")

            ws_pm = workbook.create_sheet(SHEET_NAME_PM)
            ws_pm.append(headers)
            worksheets.append(SHEET_NAME_PM)

            headers_gm = [h for h in headers if h != "Label Key"]
            ws_gm = workbook.create_sheet(SHEET_NAME_GM)
            ws_gm.append(headers_gm)
            worksheets.append(SHEET_NAME_GM)

        # Create sheets for data models
        if self.processor.is_sdm_included:
            dm_headers = ["Section", "Element/Subsection", "Field Id", "Default Label"]
            self.excel_handler.create_sheets_per_lang(
                workbook, "DataModel", xml_langs, dm_headers
            )
            for lang in xml_langs:
                worksheets.append(f"DataModel ({lang})")

        # Process each data model
        data_models = self.processor.get_all_data_models()
        is_first_pm_row = True

        for data_model in data_models:
            config_name = data_model.name
            standard_model = self.processor.get_data_model(
                f"{KEYWORD_SAP_STANDARD} {config_name}"
            )

            translatable_tags = data_model.soup.find_all(self.processor.translatable_tags)
            is_pmgm_soup = False
            translation_feature = "Manage Templates"

            # Determine worksheet and translation feature
            if data_model.model_type == DataModelType.PM_FORM_TEMPLATE:
                ws_name = SHEET_NAME_PM
                is_pmgm_soup = True
                translation_feature = "Manage Templates -> Performance Review"
            elif data_model.model_type in [
                DataModelType.GOAL_PLAN_TEMPLATE,
                DataModelType.DEVELOPMENT_PLAN_TEMPLATE
            ]:
                ws_name = SHEET_NAME_GM
                is_pmgm_soup = True
                if data_model.model_type == DataModelType.DEVELOPMENT_PLAN_TEMPLATE:
                    translation_feature = "Manage Templates -> Development"
                else:
                    translation_feature = "Manage Templates -> Goal Plan"
            else:
                ws_name = None

            ws = workbook[ws_name] if ws_name and ws_name in workbook.sheetnames else None

            prev_parent_tag = None
            prev_tag_name = None
            header_creation_map = {}
            added_countries_per_ws = {}

            for tag in translatable_tags:
                parent_tag = tag.parent
                parent_tag_name = parent_tag.name
                parent_tag_id = parent_tag.get("id")
                visibility = parent_tag.get("visibility")

                if visibility == "none" or parent_tag_name in TAGS_TO_BE_IGNORED:
                    continue

                section_name, subsection_name, country_code, skip_country = (
                    self.processor.get_section_info(tag, config_name, self.active_countries)
                )

                if skip_country:
                    continue

                default_label = self.xml_handler.get_default_title(tag, False, True)
                field = self.xml_handler.get_readable_name(tag.name, True)

                if is_pmgm_soup:
                    if parent_tag == prev_parent_tag and prev_tag_name == tag.name:
                        continue

                    if ws:
                        if remove_html_tags and "<" in default_label and "</" in default_label:
                            soup = BeautifulSoup(default_label, features="lxml")
                            default_label = soup.get_text()

                        if ws_name == SHEET_NAME_PM and is_first_pm_row:
                            row = [
                                translation_feature, config_name,
                                "General Settings", "Form Name", config_name, ""
                            ]
                            ws.append(row)
                            is_first_pm_row = False

                        row = [
                            translation_feature, config_name,
                            subsection_name, field, default_label
                        ]

                        for lang_id in xml_langs:
                            lang = lang_id.replace("-", "_")
                            lang_label_tag = parent_tag.find(
                                attrs={"lang": lang}, recursive=False
                            )

                            if lang_label_tag is None:
                                msg_key = tag.get("msgKey") or tag.get("msgkey")
                                lang_label = ""

                                if msg_key and msg_key in self.label_keys_dict:
                                    lang_label = self.label_keys_dict[msg_key].get(lang, "")

                                if ws_name == SHEET_NAME_PM and msg_key and msg_key not in row:
                                    row.append(msg_key)
                            else:
                                lang_label = lang_label_tag.string or ""

                            if remove_html_tags and "<" in lang_label and "</" in lang_label:
                                soup = BeautifulSoup(lang_label, features="lxml")
                                lang_label = soup.get_text()

                            row.append(lang_label)

                        ws.append(row)

                    prev_parent_tag = parent_tag
                    prev_tag_name = tag.name

                else:
                    # Data model translations
                    if parent_tag != prev_parent_tag:
                        missing_langs = self.xml_handler.get_missing_langs(tag, xml_langs)

                        for missing_lang in missing_langs:
                            dm_ws_name = f"DataModel ({missing_lang})"
                            if dm_ws_name not in workbook.sheetnames:
                                continue

                            dm_ws = workbook[dm_ws_name]
                            standard_label = ""

                            if standard_model:
                                std_tag = standard_model.soup.find(
                                    name=parent_tag_name, attrs={"id": parent_tag_id}
                                )
                                if std_tag:
                                    lang_tag = self.xml_handler.get_lang_tag_of(
                                        std_tag, missing_lang
                                    )
                                    if lang_tag:
                                        standard_label = lang_tag.string or ""

                            row = [
                                section_name, subsection_name,
                                parent_tag_id, default_label, standard_label
                            ]

                            if parent_tag_name in HIGHLIGHT_TAGS:
                                self.excel_handler.append_as_header_row(dm_ws, row)
                            else:
                                dm_ws.append(row)

                    prev_parent_tag = parent_tag

                    lang_attr = tag.get("xml:lang")
                    if lang_attr:
                        dm_ws_name = f"DataModel ({lang_attr})"
                        if dm_ws_name in workbook.sheetnames:
                            dm_ws = workbook[dm_ws_name]
                            label = tag.string or ""

                            row = [
                                section_name, subsection_name,
                                parent_tag_id, default_label, label
                            ]

                            if parent_tag_name in HIGHLIGHT_TAGS:
                                self.excel_handler.append_as_header_row(dm_ws, row)
                            else:
                                dm_ws.append(row)

    def save_workbook(
        self,
        workbook: Workbook,
        save_dir: str,
        filename: Optional[str] = None
    ) -> str:
        """
        Save workbook to file.

        Args:
            workbook: Workbook to save
            save_dir: Directory to save to
            filename: Optional filename (generated if not provided)

        Returns:
            Path to saved file
        """
        if not filename:
            filename = self.excel_handler.generate_export_filename(save_dir)
        else:
            filename = os.path.join(save_dir, filename)

        self.excel_handler.prepare_and_save_workbook(workbook, filename)
        return filename

    def set_label_keys(
        self,
        label_keys_dict: Dict[str, Dict],
        headers: List[str]
    ):
        """Set label keys dictionary for PM form templates."""
        self.label_keys_dict = label_keys_dict
        self.label_keys_headers = headers

    def set_active_countries(self, countries: List[str]):
        """Set active countries for CSF filtering."""
        self.active_countries = countries

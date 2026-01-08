"""
Translation Importer Module

Imports translations from Excel workbooks back to SF data models.
"""

import os
import time
from typing import List, Optional, Dict, Any, Callable, Tuple

from openpyxl import Workbook
from bs4 import BeautifulSoup

from ..config import (
    SHEET_NAME_PM,
    SHEET_NAME_GM,
    SHEET_NAME_PL,
    CHILD_CHAR
)
from ..models.datamodel import DataModel, ImportResult
from ..io.xml_handler import XMLHandler
from ..io.excel_handler import ExcelHandler
from ..io.csv_handler import CSVHandler
from .datamodel_processor import DataModelProcessor


class TranslationImporter:
    """Imports translations from Excel back to data models."""

    def __init__(
        self,
        processor: DataModelProcessor,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ):
        self.processor = processor
        self.xml_handler = XMLHandler()
        self.excel_handler = ExcelHandler()
        self.csv_handler = CSVHandler()
        self.progress_callback = progress_callback

        # State
        self.label_keys_dict: Dict[str, Dict] = {}
        self.label_keys_headers: List[str] = []
        self.import_logs: List[str] = []
        self.modified_models: List[DataModel] = []

    def _log_progress(self, percent: int, message: str):
        """Log progress if callback is set."""
        if self.progress_callback:
            self.progress_callback(percent, message)

    def _log_import(self, message: str):
        """Add message to import logs."""
        timestamp = time.strftime("%a_%d%b_%Y_%Hh%Mm%Ss", time.localtime())
        self.import_logs.append(f"{timestamp}: {message}")

    def import_from_workbook(
        self,
        workbook: Workbook,
        worksheets_to_process: List[str],
        save_dir: str
    ) -> ImportResult:
        """
        Import translations from workbook to data models.

        Args:
            workbook: Translations workbook
            worksheets_to_process: List of worksheet names to process
            save_dir: Directory to save output files

        Returns:
            ImportResult with operation details
        """
        self.import_logs = []
        self.modified_models = []
        progress = 0
        progress_incr = 55 / len(worksheets_to_process) if worksheets_to_process else 0

        result = ImportResult(success=True)

        for ws_name in worksheets_to_process:
            progress += progress_incr
            self._log_progress(
                int(progress),
                f"Processing '{ws_name}' sheet from workbook..."
            )

            ws = workbook[ws_name]
            ws.protection.password = "...ApTrans..."
            ws.protection.disable()

            max_col = self._get_max_column(ws)
            if max_col < 2:
                continue

            # Add change log column
            change_log_col = self.excel_handler.add_change_log_column(ws)

            if ws_name.startswith("DataModel"):
                self._process_datamodel_sheet(ws, ws_name, change_log_col)
            elif ws_name in [SHEET_NAME_PM, SHEET_NAME_GM]:
                self._process_pmgm_sheet(ws, ws_name, change_log_col, save_dir)

            ws.protection.enable()
            ws.protection.sort = False
            ws.protection.autoFilter = False
            ws.protection.formatColumns = False

            result.changes_made += 1

        # Save modified workbook with change log
        if self.modified_models:
            progress += 5
            self._log_progress(progress, "Saving updated workbook with change log...")

            workbook_path = os.path.join(save_dir, "TranslationsWorkbook_WithChangeLog.xlsx")
            workbook.save(workbook_path)

        # Save import log
        progress += 5
        log_path = self._save_import_log(save_dir)
        result.log_file_path = log_path

        # Generate ready-to-import XML files
        progress_incr = 35 / len(self.modified_models) if self.modified_models else 0

        for model in self.modified_models:
            progress += progress_incr
            file_name = f"ReadyToImport_{model.name}.xml"
            file_path = os.path.join(save_dir, file_name)

            self.xml_handler.write_xml_file(model.soup, file_path)
            result.files_generated.append(file_path)

            self._log_progress(
                int(progress),
                f"Generated ready-to-import file: {file_name}"
            )

        self._log_progress(100, "Import complete!")
        return result

    def _get_max_column(self, ws) -> int:
        """Get the maximum non-empty column number."""
        col_num = 1
        max_col = ws.max_column
        for cell in ws[1]:
            if cell.value is None or cell.value == "":
                max_col = col_num
                break
            col_num += 1
        return max_col

    def _process_datamodel_sheet(
        self,
        ws,
        ws_name: str,
        change_log_col: int
    ):
        """Process a DataModel worksheet."""
        # Extract language from sheet name
        lang_id = ws_name[11:-1]  # "DataModel (xx-XX)" -> "xx-XX"

        row_num = 1
        parent_tag = None
        grand_parent_tag = None

        while row_num < ws.max_row:
            row_num += 1

            dm_ref = ws.cell(row=row_num, column=1).value
            translatable_item = ws.cell(row=row_num, column=2).value
            tag_id = ws.cell(row=row_num, column=3).value
            lang_label = ws.cell(row=row_num, column=5).value

            if not dm_ref:
                break

            # Handle CSF data models
            is_csf = False
            if "(" in dm_ref:
                dm_name = dm_ref[:dm_ref.find("(") - 1].strip()
                dm_ref = dm_name
                is_csf = True

            if dm_ref == "Employee Profile":
                dm_ref = "SFEC Succession Data Model"

            data_model = self.processor.get_data_model(dm_ref)
            if not data_model:
                ws.cell(row=row_num, column=change_log_col).value = (
                    f"No data model found for {dm_ref}"
                )
                continue

            soup = data_model.soup

            # Handle header rows
            col1_bold = ws.cell(row=row_num, column=1).font.b
            col2_bold = ws.cell(row=row_num, column=2).font.b
            col3_bold = ws.cell(row=row_num, column=3).font.b

            if col1_bold and col2_bold and col3_bold:
                if translatable_item == "country" or grand_parent_tag is None:
                    grand_parent_tag = soup
                elif grand_parent_tag:
                    grand_parent_tag = parent_tag

                parent_tag = grand_parent_tag.find(
                    name=translatable_item,
                    attrs={"id": tag_id, "visibility": "both"}
                )
                if parent_tag is None:
                    parent_tag = soup.find(
                        name=translatable_item, attrs={"id": tag_id}
                    )

            if parent_tag is None:
                parent_tag = soup

            # Find matching tag
            matching_tag = parent_tag.find(
                name=translatable_item,
                attrs={"id": tag_id, "visibility": "both"}
            )
            if matching_tag is None:
                matching_tag = parent_tag.find(
                    name=translatable_item, attrs={"id": tag_id}
                )
                if matching_tag is None:
                    matching_tag = soup.find(
                        name=translatable_item, attrs={"id": tag_id}
                    )

            if matching_tag is None:
                self._log_import(
                    f"No matching tag found in {dm_ref} for {translatable_item} ({tag_id})"
                )
                ws.cell(row=row_num, column=change_log_col).value = (
                    f"No matching tag found in {dm_ref}"
                )
                continue

            # Find or create label tag
            label_tag_name = "label"
            matching_label = matching_tag.find(
                name=label_tag_name, attrs={"xml:lang": lang_id}
            )

            if matching_tag.find(label_tag_name) is None:
                label_tag_name = "instruction"
                matching_label = matching_tag.find(
                    name=label_tag_name, attrs={"xml:lang": lang_id}
                )
                if matching_tag.find(label_tag_name) is None:
                    matching_label = None
                    label_tag_name = "label"

            if matching_label is None:
                # Create new label tag
                if lang_label:
                    new_label = soup.new_tag(label_tag_name)
                    new_label["xml:lang"] = lang_id
                    new_label.string = lang_label
                    matching_tag.insert(2, new_label)

                    if data_model not in self.modified_models:
                        self.modified_models.append(data_model)

                    self._log_import(
                        f"Row {row_num}: Added '{lang_id}' translation for {translatable_item}"
                    )
                    ws.cell(row=row_num, column=change_log_col).value = (
                        f"Translation Added: '{lang_label}'"
                    )
            else:
                # Update existing label
                old_label = matching_label.string
                if lang_label != old_label:
                    if data_model not in self.modified_models:
                        self.modified_models.append(data_model)

                    if lang_label:
                        matching_label.string = lang_label
                        self._log_import(
                            f"Row {row_num}: Changed '{lang_id}' translation for "
                            f"{translatable_item} from '{old_label}' to '{lang_label}'"
                        )
                        ws.cell(row=row_num, column=change_log_col).value = (
                            f"Translation Changed from '{old_label}' to '{lang_label}'"
                        )

    def _process_pmgm_sheet(
        self,
        ws,
        ws_name: str,
        change_log_col: int,
        save_dir: str
    ):
        """Process a PM/GM template worksheet."""
        parent_tag_name = "sf-form" if ws_name == SHEET_NAME_PM else "obj-plan-template"

        # Prepare label key rows for writing
        label_key_rows = []
        default_lang = None
        for key in self.label_keys_dict:
            if default_lang is None:
                default_lang = self.label_keys_dict[key].get("default")
            label_key_rows.append(self.label_keys_dict[key])

        # Extract language columns
        lang_keys = []
        lang_start_col = 0
        col_num = 1

        while col_num < ws.max_column:
            header = ws.cell(row=1, column=col_num).value
            if header and "(" in header:
                if lang_start_col == 0:
                    lang_start_col = col_num
                lang_key = header[header.rfind("(") + 1:header.rfind(")")]
                lang_keys.append(lang_key)
            col_num += 1

        row_num = 1
        max_col = self._get_max_column(ws)
        parent_section_tag = None
        parent_tag = None
        unmatched_rows = []
        unmatched_templates = []

        while row_num < ws.max_row:
            row_num += 1

            template_name = ws.cell(row=row_num, column=2).value
            if template_name is None:
                break

            section_name = ws.cell(row=row_num, column=3).value
            translatable_item = ws.cell(row=row_num, column=4).value
            change_text = None

            data_model = self.processor.get_data_model(template_name)
            if not data_model:
                unmatched_rows.append(row_num)
                if template_name not in unmatched_templates:
                    unmatched_templates.append(template_name)
                continue

            soup = data_model.soup

            # Parse section name for attributes
            if section_name and "(" in section_name:
                attr_name = section_name[section_name.rfind("(") + 1:section_name.rfind("=")]
                attr_value = section_name[section_name.rfind("=") + 1:section_name.rfind(")")]
                section_name_clean = section_name[:section_name.find("(")]
                section_tag_name = self.xml_handler.derive_section_tag_name(section_name_clean)

                parent_section_tag = soup.find(
                    section_tag_name, attrs={attr_name: attr_value}
                )
                if parent_section_tag is None:
                    parent_section_tag = soup.find(
                        "fm-sect", attrs={attr_name: attr_value}
                    )
                parent_tag = parent_section_tag
            elif section_name == f"{CHILD_CHAR}Section Configuration":
                if parent_section_tag:
                    parent_tag = parent_section_tag.find("fm-sect-config")
            else:
                parent_tag = soup.find(parent_tag_name)

            # Extract field name from translatable item
            if translatable_item and "(" in translatable_item:
                translatable_item = translatable_item[
                    translatable_item.rfind("(") + 1:translatable_item.rfind(")")
                ]

            if parent_tag is None:
                continue

            tag = parent_tag.find(translatable_item)
            if tag is None:
                continue

            # Collect language labels from worksheet
            lang_labels = {}
            for n in range(lang_start_col, max_col):
                lang_val = ws.cell(row=row_num, column=n).value
                if n - lang_start_col < len(lang_keys):
                    lang_key = lang_keys[n - lang_start_col]
                    lang_labels[lang_key] = lang_val

            # Process tag
            def_label = ws.cell(row=row_num, column=5).value
            msg_key_from_xl = ws.cell(row=row_num, column=6).value if ws_name == SHEET_NAME_PM else None
            tag_name = tag.name
            modified_langs = []
            old_labels = []
            new_labels = []

            if ws_name == SHEET_NAME_PM:
                change_text = self._process_pm_tag(
                    tag, soup, data_model, tag_name, lang_keys, lang_labels,
                    label_key_rows, modified_langs, old_labels, new_labels,
                    row_num, ws_name
                )
            else:
                change_text = self._process_gm_tag(
                    tag, soup, data_model, tag_name, lang_keys, lang_labels,
                    def_label, modified_langs, old_labels, new_labels,
                    row_num, ws_name
                )

            if change_text:
                ws.cell(row=row_num, column=change_log_col).value = change_text

        # Write updated label keys file for PM
        if ws_name == SHEET_NAME_PM and label_key_rows:
            label_keys_path = os.path.join(save_dir, "ReadyToImport_FormLabelKeys.csv")
            self.csv_handler.write_csv_from_dict_list(
                label_keys_path, label_key_rows, self.label_keys_headers
            )
            self._log_progress(0, f"Generated ReadyToImport_FormLabelKeys.csv")

    def _process_pm_tag(
        self,
        tag,
        soup,
        data_model,
        tag_name: str,
        lang_keys: List[str],
        lang_labels: Dict[str, str],
        label_key_rows: List[Dict],
        modified_langs: List[str],
        old_labels: List[str],
        new_labels: List[str],
        row_num: int,
        ws_name: str
    ) -> Optional[str]:
        """Process a PM template tag."""
        change_text = None

        if tag.has_attr("msgKey") or tag.has_attr("msgkey"):
            lowercase = False
            msg_key = tag.get("msgKey")
            if msg_key is None:
                msg_key = tag.get("msgkey")
                lowercase = True

            label_key_dict = self.label_keys_dict.get(msg_key)

            if label_key_dict:
                for lang_key in lang_keys:
                    label_from_csv = label_key_dict.get(lang_key)
                    label_from_xl = lang_labels.get(lang_key)

                    if label_from_csv and label_from_csv != label_from_xl:
                        for lk_row in label_key_rows:
                            if lk_row.get("label_key") == msg_key:
                                lk_row[lang_key] = label_from_xl
                                modified_langs.append(lang_key)
                                old_labels.append(label_from_csv)
                                new_labels.append(label_from_xl)
                                break

                if modified_langs:
                    self._log_import(
                        f"Row {row_num}: Updated FormLabelKeys for '{msg_key}' "
                        f"languages {modified_langs}"
                    )
                    change_text = (
                        f"Translation changed in FormLabelKeys for "
                        f"{modified_langs} to {new_labels}"
                    )

                # Fix lowercase msgkey
                if lowercase:
                    del tag["msgkey"]
                    tag["msgKey"] = msg_key
                    if data_model not in self.modified_models:
                        self.modified_models.append(data_model)

        return change_text

    def _process_gm_tag(
        self,
        tag,
        soup,
        data_model,
        tag_name: str,
        lang_keys: List[str],
        lang_labels: Dict[str, str],
        def_label: str,
        modified_langs: List[str],
        old_labels: List[str],
        new_labels: List[str],
        row_num: int,
        ws_name: str
    ) -> Optional[str]:
        """Process a GM template tag."""
        change_text = None

        # Update default label
        if not tag.has_attr("lang"):
            def_val_from_xml = tag.string
            if def_label and def_val_from_xml != def_label:
                tag.string = f"<![CDATA[{def_label}]]>"
                modified_langs.append("Default")
                old_labels.append(def_val_from_xml)
                new_labels.append(def_label)

        # Update language-specific labels
        next_siblings = tag.find_next_siblings(tag_name)
        matching_tag = tag

        for lang_key in lang_labels:
            lang_val = lang_labels.get(lang_key)
            if not lang_val or not lang_val.strip():
                continue

            lang_found = False
            for sibling in next_siblings:
                if sibling.get("lang") == lang_key:
                    matching_tag = sibling
                    lang_found = True
                    old_val = sibling.string
                    if old_val != lang_val:
                        sibling.string = f"<![CDATA[{lang_val}]]>"
                        modified_langs.append(lang_key)
                        old_labels.append(old_val)
                        new_labels.append(lang_val)
                    break

            if not lang_found:
                # Create new language tag
                new_lang_tag = soup.new_tag(tag_name)
                new_lang_tag["lang"] = lang_key
                new_lang_tag.string = f"<![CDATA[{lang_val}]]>"
                matching_tag.insert_after(new_lang_tag)

                if data_model not in self.modified_models:
                    self.modified_models.append(data_model)

        if modified_langs:
            if data_model not in self.modified_models:
                self.modified_models.append(data_model)

            self._log_import(
                f"Row {row_num}: Changed translations for {modified_langs}"
            )
            change_text = (
                f"Translation Changed for {modified_langs} "
                f"from {old_labels} to {new_labels}"
            )

        return change_text

    def _save_import_log(self, save_dir: str) -> str:
        """Save import log to file."""
        timestamp = time.strftime("%a_%d%b_%Y_%Hh%Mm%Ss", time.localtime())
        log_file = os.path.join(save_dir, f"ImportLog_{timestamp}.log")

        with open(log_file, "wt", encoding="utf-8") as f:
            f.write("\n\n".join(self.import_logs))

        return log_file

    def set_label_keys(
        self,
        label_keys_dict: Dict[str, Dict],
        headers: List[str]
    ):
        """Set label keys dictionary for PM form templates."""
        self.label_keys_dict = label_keys_dict
        self.label_keys_headers = headers
